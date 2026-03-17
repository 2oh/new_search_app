import tempfile
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# ✅ main.py から関数を直接import
import sys
from pathlib import Path

# プロジェクト直下（main.py がある場所）を import パスに追加
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from main import detect_columns, extract_data_from_excel, get_quantity_colors


def make_excel(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # ヘッダは3行目（header_row_index=2）
    ws["A1"] = "タイトル"
    ws["A2"] = "サブ"
    ws.append(["品番", "PG名", "品名", "数量"])   # row=3

    ws.append(["A001", "PG01", "部品1", 1])      # row=4（色付き）
    ws.append(["A002", "PG02", "部品2", None])   # row=5（数量空→フィルタで落ちる）
    ws.append(["A003", "PG03", "部品3", 0])      # row=6（残る）

    # D4（数量セル）に色を付ける
    ws["D4"].fill = PatternFill(patternType="solid", fgColor="FFFF00")

    wb.save(path)


def test_quantity_color_alignment():
    with tempfile.TemporaryDirectory() as td:
        xlsx = Path(td) / "t.xlsx"
        make_excel(xlsx)

        sheet = "Sheet1"
        header_row_index = 2

        detected = detect_columns(str(xlsx), sheet, header_row_index)
        df = extract_data_from_excel(str(xlsx), sheet, detected, header_row_index)

        # 数量空(A002)が落ちるので A001, A003 の2行になる
        assert list(df["品番"]) == ["A001", "A003"]

        qty_colors = get_quantity_colors(str(xlsx), sheet, header_row_index, detected["数量"])

        # ✅ わざと「修正前のバグ」を再現する割り当て（先頭から切る）
        df["数量セル色"] = qty_colors[: len(df)]

        # A001は色あり、A003は色なし…のはず（ここがズレると落ちる）
        assert df.loc[df["品番"] == "A001", "数量セル色"].iloc[0] != ""
        assert df.loc[df["品番"] == "A003", "数量セル色"].iloc[0] == ""