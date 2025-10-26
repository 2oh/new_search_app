# ======================================================
#  Excel–PDF結合アプリ（フェーズ2.7.1）
#  機能: フォルダボタン復活＋初期フォルダconfig対応
# ======================================================

import flet as ft
import pandas as pd
import os
import json
import re
from openpyxl import load_workbook

CONFIG_FILE = "config.json"

# ===== normalize() =====
def normalize(text: str) -> str:
    """列名比較用の正規化関数"""
    if text is None:
        return ""
    s = str(text).strip().lower()
    s = s.translate(str.maketrans(
        "０１２３４５６７８９ａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐｑｒｓｔｕｖｗｘｙｚ",
        "0123456789abcdefghijklmnopqrstuvwxyz"
    ))
    s = re.sub(r"[／/・\.\s　\-＿_]", "", s)
    return s


# ===== detect_columns() =====
def detect_columns(file_path: str, sheet_name: str, header_row_index: int, target_columns=None):
    if target_columns is None:
        target_columns = ["品番", "PG名", "品名", "数量"]

    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row_index)
    detected = {}
    normalized_cols = [normalize(c) for c in df.columns]

    for target in target_columns:
        norm_target = normalize(target)
        for i, col in enumerate(normalized_cols):
            if norm_target in col or col in norm_target:
                detected[target] = i
                break
            if norm_target.startswith("pg") and ("pg" in col or "ｐｇ" in col):
                detected[target] = i
                break
    print(f"[DEBUG] detect_columns detected={detected}")
    return detected


# ===== Excel抽出関数 =====
def extract_data_from_excel(file_path: str, sheet_name: str, detected_columns: dict, header_row_index: int):
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row_index)

    col_items = []
    for key in ["品番", "PG名", "品名", "数量"]:
        if key in detected_columns:
            col_name = df.columns[detected_columns[key]]
            col_items.append(col_name)

    sub_df = df[col_items].copy() if col_items else pd.DataFrame()
    cols = list(df.columns)

    # --- 品名結合 ---
    def find_name_block_indices(cols):
        for i, c in enumerate(cols):
            if "品名" in str(c):
                j = i + 1
                block = [i]
                while j < len(cols):
                    cj = str(cols[j])
                    if cj.startswith("Unnamed") or cj.strip() == "" or cj.lower() in ("none", "nan"):
                        block.append(j)
                        j += 1
                    else:
                        break
                return block
        return []

    name_idx = find_name_block_indices(cols)
    if name_idx:
        name_df = df.iloc[:, name_idx].astype(str)
        sub_df["品名"] = (
            name_df.apply(
                lambda row: " ".join(v.strip() for v in row if v.strip() and v.lower() != "nan"),
                axis=1
            ).str.strip()
        )
    else:
        name_like_cols = [c for c in df.columns if "品名" in str(c)]
        if len(name_like_cols) == 1:
            sub_df["品名"] = df[name_like_cols[0]].astype(str).fillna("").str.strip()

    # --- 数量フィルタ ---
    quantity_col = next((c for c in df.columns if "数量" in str(c)), None)
    if quantity_col:
        def keep_row(x):
            if pd.isna(x): return False
            s = str(x).strip()
            if s == "": return False
            try:
                return float(s.replace(",", "")) != 0.0
            except ValueError:
                return True
        sub_df = sub_df[df[quantity_col].apply(keep_row)]

    return sub_df


# ===== 背景色検出 =====
def get_quantity_colors(file_path, sheet_name, header_row_index, quantity_col_index):
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    color_list = []
    for i, row in enumerate(ws.iter_rows(min_row=header_row_index + 2)):
        if quantity_col_index < len(row):
            cell = row[quantity_col_index]
            fill = cell.fill
            if fill and fill.start_color and fill.start_color.rgb and fill.start_color.rgb != "00000000":
                color_list.append("あり")
            else:
                color_list.append("")
    return color_list


# ===== detect_header_row() =====
def detect_header_row(df, merged_cells_info):
    merged_a_rows = set()
    for crange in merged_cells_info:
        if crange.min_col == 1:
            for r in range(crange.min_row - 1, crange.max_row):
                merged_a_rows.add(r)

    print("=== detect_header_row: start ===")
    for i, row in df.iterrows():
        a_val = str(row.iloc[0]).strip() if not pd.isna(row.iloc[0]) else ""
        filled = sum(1 for v in row if not pd.isna(v) and str(v).strip() != "")
        print(f"Row {i+1}: A='{a_val}', filled={filled}, mergedA={i in merged_a_rows}")

        if i in merged_a_rows:
            continue
        if not a_val:
            continue
        if filled >= 1:
            print(f"✅ detected header_row_index = {i}")
            print("===============================")
            return i

    print("⚠️ detect_header_row: fallback to row 0")
    print("===============================")
    return 0


# ===== メインUI =====
def main(page: ft.Page):
    page.title = "Excel–PDF結合アプリ（フェーズ2.7.1）"
    page.scroll = "adaptive"

    # ------------------------------
    # 設定ファイルの読み込み
    # ------------------------------
    def load_config():
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        return {
            "excel_folder": "",
            "pdf_folder": "",
            "output_folder": "",
            "search_mode": "Excel",
            "target_columns": ["品番", "PG名"]
        }

    def save_config(e):
        config.update({
            "excel_folder": excel_folder_field.value,
            "pdf_folder": search_folder_field.value,
            "output_folder": output_folder_field.value,
            "search_mode": mode_dropdown.value,
            "target_columns": [x.strip() for x in target_col_field.value.split(",") if x.strip()]
        })
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(config, f, indent=2, ensure_ascii=False)

    config = load_config()

    # ------------------------------
    # UI 要素
    # ------------------------------
    excel_folder_field = ft.TextField(label="選択中のExcelファイル", expand=True)
    search_folder_field = ft.TextField(label="検索先フォルダ（PDF）", value=config.get("pdf_folder", ""), expand=True)
    output_folder_field = ft.TextField(label="出力先フォルダ", value=config.get("output_folder", ""), expand=True)

    # --- 検索モードUI ---
    mode_dropdown = ft.Dropdown(
        label="検索モード",
        options=[ft.dropdown.Option("Excel"), ft.dropdown.Option("通常")],
        value=config.get("search_mode", "Excel"),
        width=150
    )

    target_col_field = ft.TextField(
        label="抽出対象列（Excelモード）",
        value=", ".join(config.get("target_columns", [])),
        width=400,
        visible=(config.get("search_mode", "Excel") == "Excel")
    )

    manual_keyword_field = ft.TextField(
        label="検索文字列（通常モード）",
        value="",
        width=400,
        visible=(config.get("search_mode", "Excel") != "Excel")
    )

    # --- 検索モード切替処理 ---
    def on_mode_change(e):
        if mode_dropdown.value == "Excel":
            target_col_field.visible = True
            manual_keyword_field.visible = False
        else:
            target_col_field.visible = False
            manual_keyword_field.visible = True
        page.update()

    mode_dropdown.on_change = on_mode_change

    # --- 検索モード行（左：入力欄群 / 右：保存ボタン） ---
    config_row = ft.Row(
        [
            ft.Row([mode_dropdown, target_col_field, manual_keyword_field], alignment="start"),
            ft.ElevatedButton("設定を保存", on_click=save_config)
        ],
        alignment="spaceBetween",
    )

    # --- フォルダ選択イベント ---
    folder_picker_search = ft.FilePicker(on_result=lambda e: on_folder_picked(e, "search"))
    folder_picker_output = ft.FilePicker(on_result=lambda e: on_folder_picked(e, "output"))
    page.overlay.append(folder_picker_search)
    page.overlay.append(folder_picker_output)

    def on_folder_picked(e, mode):
        if not e.path:
            return
        if mode == "search":
            search_folder_field.value = e.path
            config["pdf_folder"] = e.path
        elif mode == "output":
            output_folder_field.value = e.path
            config["output_folder"] = e.path
        page.update()

    # ✅ 初期フォルダを config に合わせて開くよう修正
    def pick_search_folder(e):
        start_path = config.get("pdf_folder", "")
        if not os.path.isdir(start_path):
            start_path = os.getcwd()
        folder_picker_search.get_directory_path(initial_directory=start_path)

    def pick_output_folder(e):
        start_path = config.get("output_folder", "")
        if not os.path.isdir(start_path):
            start_path = os.getcwd()
        folder_picker_output.get_directory_path(initial_directory=start_path)

    # --- Excel関連UI ---
    selected_excel_path = ""
    sheet_dropdown = ft.Dropdown(label="シート選択", width=300)
    message = ft.Text("")
    table = ft.DataTable(columns=[ft.DataColumn(ft.Text("項目"))], rows=[])

    # ------------------------------
    # Excelヘルパー関数
    # ------------------------------
    def get_merged_cells_info(excel_path, sheet_name):
        wb = load_workbook(excel_path, data_only=True)
        ws = wb[sheet_name]
        return ws.merged_cells.ranges

    # ------------------------------
    # イベント処理
    # ------------------------------
    def pick_excel_result(e: ft.FilePickerResultEvent):
        nonlocal selected_excel_path
        if not e.files:
            return
        selected_excel_path = e.files[0].path
        excel_folder_field.value = os.path.abspath(selected_excel_path)
        message.value = ""
        table.rows.clear()
        table.columns = [ft.DataColumn(ft.Text("項目"))]
        page.update()

        try:
            xls = pd.ExcelFile(selected_excel_path)
            sheet_dropdown.options = [ft.dropdown.Option(name) for name in xls.sheet_names]
            sheet_dropdown.value = None
            sheet_dropdown.on_change = on_sheet_selected
            page.snack_bar = ft.SnackBar(ft.Text(f"{len(xls.sheet_names)} シートを読み込みました。選択してください。"))
            page.snack_bar.open = True
            page.update()
        except PermissionError:
            page.dialog = ft.AlertDialog(
                title=ft.Text("ファイル使用中エラー"),
                content=ft.Text("Excelファイルを閉じてから再実行してください。")
            )
            page.dialog.open = True
            page.update()

    def on_sheet_selected(e):
        message.value = ""
        table.rows.clear()
        page.update()

    def on_extract_click(e):
        if not selected_excel_path or not sheet_dropdown.value:
            message.value = "Excelファイルとシートを選択してください。"
            page.update()
            return

        try:
            df_raw = pd.read_excel(selected_excel_path, sheet_name=sheet_dropdown.value, header=None)
            merged_info = get_merged_cells_info(selected_excel_path, sheet_dropdown.value)
            header_row_index = detect_header_row(df_raw, merged_info)
            detected_columns = detect_columns(selected_excel_path, sheet_dropdown.value, header_row_index)
            df = extract_data_from_excel(selected_excel_path, sheet_dropdown.value, detected_columns, header_row_index)

            # 背景色列追加
            if "数量" in detected_columns:
                color_list = get_quantity_colors(
                    selected_excel_path,
                    sheet_dropdown.value,
                    header_row_index,
                    detected_columns["数量"]
                )
                df["数量セル色"] = color_list[:len(df)]

            if df.empty:
                message.value = "抽出結果がありません。"
                table.columns = [ft.DataColumn(ft.Text("項目"))]
                table.rows = []
            else:
                message.value = f"{len(df)}件のデータを抽出しました。"
                table.columns = [ft.DataColumn(ft.Text(c)) for c in df.columns]
                table.rows = [
                    ft.DataRow(cells=[ft.DataCell(ft.Text(str(v))) for v in row])
                    for row in df.values.tolist()
                ]
            page.update()
        except Exception as ex:
            message.value = f"エラー: {ex}"
            page.update()

    # --- ファイル選択 ---
    file_picker = ft.FilePicker(on_result=pick_excel_result)
    page.overlay.append(file_picker)

    def pick_excel_click(e):
        file_picker.pick_files(allowed_extensions=["xlsx", "xls"])

    # ------------------------------
    # レイアウト構成
    # ------------------------------
    layout = ft.Column(
        [
            ft.Text("⚙️ 設定", size=20, weight=ft.FontWeight.BOLD),
            config_row,
            ft.Column(
                [
                    ft.Row([search_folder_field, ft.ElevatedButton("検索先を選択", on_click=pick_search_folder)], alignment="center"),
                    ft.Row([output_folder_field, ft.ElevatedButton("出力先を選択", on_click=pick_output_folder)], alignment="center"),
                ],
                spacing=10,
            ),
            ft.Divider(),
            ft.Text("🔍 Excel検索モード", size=20, weight=ft.FontWeight.BOLD),
            ft.Row([excel_folder_field, ft.ElevatedButton("Excelを選択", on_click=pick_excel_click)], alignment="center"),
            ft.Row([sheet_dropdown, ft.ElevatedButton("抽出実行", on_click=on_extract_click)], alignment="center"),
            ft.Divider(),
            message,
            table
        ],
        expand=True,
        scroll="auto"
    )

    page.add(layout)


ft.app(target=main)
