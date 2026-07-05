# ======================================================
#  図面PDF検索結合ツール
# ======================================================

import flet as ft
import pandas as pd
import os
import json
import re
import sys
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

CONFIG_FILE = "config.json"
PREVIEW_CACHE_DIR = "preview_cache"
APP_ICON_FILE = "assets/app_icon.ico"
APP_VERSION = "v1.0.6"
DEBUG = False

PDF_FILE_NAME_COLOR = ft.Colors.BLUE_200
PDF_FILE_NAME_HOVER_COLOR = ft.Colors.CYAN_100

OUTPUT_PDF_COLUMN_WIDTH = 260

def debug_print(*args, **kwargs):
    if DEBUG:
        print(*args, **kwargs)

# ========= 共通ユーティリティ =========
def normalize(text: str) -> str:
    if text is None:
        return ""
    s = str(text).strip().lower()
    s = s.translate(str.maketrans(
        "０１２３４５６７８９ａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐｑｒｓｔｕｖｗｘｙｚ",
        "0123456789abcdefghijklmnopqrstuvwxyz"
    ))
    s = re.sub(r"[／/⁄・\.\s　\-＿_]", "", s)
    return s

def normalize_filename_match(text: str) -> str:
    """PDFファイル名検索用の軽い正規化"""
    if text is None:
        return ""

    s = str(text).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s

def resource_path(relative_path: str) -> Path:
    """
    通常実行時とPyInstaller実行時の両方で、
    同梱ファイルのパスを取得する。
    """
    if hasattr(sys, "_MEIPASS"):
        return Path(sys._MEIPASS) / relative_path

    return Path(relative_path)

def collect_pdf_files(root_folder: str, progress_callback=None) -> list[Path]:
    """
    指定フォルダ配下のPDFを再帰的に集める。

    NAS上では Path.rglob("*") + is_file() が重くなることがあるため、
    os.walk でファイル一覧を取得する。
    """
    if not root_folder or not os.path.isdir(root_folder):
        return []

    pdf_files = []

    for dirpath, dirnames, filenames in os.walk(root_folder):
        for filename in filenames:
            if filename.lower().endswith(".pdf"):
                pdf_files.append(Path(dirpath) / filename)

                if progress_callback and len(pdf_files) % 1000 == 0:
                    progress_callback(len(pdf_files))

    return pdf_files

def find_pdf_candidates_by_filename(search_text: str, pdf_files: list[Path]) -> list[str]:
    """検索文字列を使ってPDFファイル名を部分一致検索する"""
    needle = normalize_filename_match(search_text)

    if not needle:
        return []

    hits = []

    for pdf_path in pdf_files:
        stem_norm = normalize_filename_match(pdf_path.stem)
        name_norm = normalize_filename_match(pdf_path.name)

        if needle in stem_norm or needle in name_norm:
            hits.append(str(pdf_path))

    return hits

def build_pdf_search_index(pdf_files: list[Path]) -> list[tuple[str, str, str]]:
    """
    PDF検索用の簡易インデックスを作る。

    Returns
    -------
    list[tuple[str, str, str]]
        (PDFフルパス, 正規化済みstem, 正規化済みファイル名)
    """
    index = []

    for pdf_path in pdf_files:
        path_text = str(pdf_path)
        stem_norm = normalize_filename_match(pdf_path.stem)
        name_norm = normalize_filename_match(pdf_path.name)
        index.append((path_text, stem_norm, name_norm))

    return index


def find_pdf_candidates_from_index(
    search_text: str,
    pdf_index: list[tuple[str, str, str]]
) -> list[str]:
    """
    検索文字列を使って、PDF検索用インデックスから候補を探す。
    """
    needle = normalize_filename_match(search_text)

    if not needle:
        return []

    hits = []

    for path_text, stem_norm, name_norm in pdf_index:
        if needle in stem_norm or needle in name_norm:
            hits.append(path_text)

    return hits

def merge_pdfs(pdf_paths: list[str], output_path: str, progress_callback=None) -> str:
    """
    複数のPDFを1つに結合して保存する。

    Parameters
    ----------
    pdf_paths:
        結合するPDFファイルのパス一覧。
        この順番で結合される。

    output_path:
        出力PDFの保存先パス。

    Returns
    -------
    str
        作成されたPDFファイルのパス。

    Raises
    ------
    ValueError:
        pdf_paths が空、または output_path が空の場合。

    FileNotFoundError:
        入力PDFが存在しない場合。
    """
    if not pdf_paths:
        raise ValueError("結合対象のPDFがありません。")

    if not output_path or not str(output_path).strip():
        raise ValueError("出力先PDFパスが指定されていません。")

    # pypdf が未インストールの場合でも、アプリ起動時には落ちないように関数内でimportする
    try:
        from pypdf import PdfReader, PdfWriter
    except ImportError as ex:
        raise ImportError(
            "pypdf がインストールされていません。'pip install pypdf' を実行してください。"
        ) from ex

    writer = PdfWriter()

    total = len(pdf_paths)

    for i, pdf_path in enumerate(pdf_paths, start=1):
        if progress_callback:
            progress_callback(i, total, pdf_path)

        path = Path(pdf_path)

        if not path.is_file():
            raise FileNotFoundError(f"PDFファイルが見つかりません: {pdf_path}")

        reader = PdfReader(str(path))

        # パスワードなしで開ける暗号化PDFなら試す
        if reader.is_encrypted:
            try:
                reader.decrypt("")
            except Exception:
                raise ValueError(f"暗号化されたPDFを開けません: {pdf_path}")

        for page in reader.pages:
            writer.add_page(page)

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    with open(output, "wb") as f:
        writer.write(f)

    return str(output)

def clear_preview_cache(cache_dir: str = PREVIEW_CACHE_DIR):
    """
    PDFプレビューキャッシュを全削除する。
    アプリ起動時に呼び出し、preview_cache が増え続けるのを防ぐ。
    """
    cache_path = Path(cache_dir)

    if not cache_path.exists():
        return

    for image_file in cache_path.glob("preview_*.png"):
        try:
            image_file.unlink()
        except Exception:
            # 削除できないファイルがあってもアプリ本体は止めない
            pass

def create_pdf_preview_image(
    pdf_path: str,
    page_number: int = 0,
    dpi: int = 120,
    cache_dir: str = PREVIEW_CACHE_DIR,
) -> str:
    """
    PDFの指定ページをPNG画像として保存し、その画像パスを返す。

    Parameters
    ----------
    pdf_path:
        プレビュー対象のPDFファイルパス。

    page_number:
        0始まりのページ番号。通常は1ページ目なので0。

    dpi:
        画像化する際の解像度。大きいほどきれいだが重くなる。

    cache_dir:
        プレビュー画像の保存先フォルダ。

    Returns
    -------
    str
        作成されたPNG画像ファイルのパス。

    Raises
    ------
    ValueError:
        pdf_path が空、またはページ番号が範囲外の場合。

    FileNotFoundError:
        PDFファイルが存在しない場合。
    """
    if not pdf_path or not str(pdf_path).strip():
        raise ValueError("PDFパスが指定されていません。")

    pdf_file = Path(str(pdf_path).strip())

    if not pdf_file.is_file():
        raise FileNotFoundError(f"PDFファイルが見つかりません: {pdf_path}")

    try:
        import pymupdf
    except ImportError:
        try:
            import fitz as pymupdf
        except ImportError as ex:
            raise ImportError(
                "PyMuPDF がインストールされていません。"
                "'python -m pip install PyMuPDF' を実行してください。"
            ) from ex

    cache_path = Path(cache_dir)
    cache_path.mkdir(parents=True, exist_ok=True)

    # ファイルパス・更新日時・ページ番号・dpi からキャッシュ名を作る
    stat = pdf_file.stat()
    cache_key_src = f"{pdf_file.resolve()}_{stat.st_mtime_ns}_{page_number}_{dpi}"
    cache_key = str(abs(hash(cache_key_src)))
    output_image = cache_path / f"preview_{cache_key}.png"

    # すでに作成済みなら再利用
    if output_image.is_file():
        return str(output_image)

    doc = None
    try:
        doc = pymupdf.open(str(pdf_file))

        if doc.page_count == 0:
            raise ValueError(f"PDFにページがありません: {pdf_path}")

        if page_number < 0 or page_number >= doc.page_count:
            raise ValueError(
                f"ページ番号が範囲外です: {page_number + 1} / {doc.page_count}"
            )

        page = doc[page_number]
        pix = page.get_pixmap(dpi=dpi)
        pix.save(str(output_image))

        return str(output_image)

    finally:
        if doc is not None:
            doc.close()

# ========= Excel列検出 =========
def detect_columns(file_path: str, sheet_name: str, header_row_index: int, target_columns=None):
    if target_columns is None:
        target_columns = ["品番", "PG名", "品名", "材料", "数量"]

    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row_index)

    debug_print(f"\n[DEBUG] ===== detect_columns start =====")
    debug_print(f"[DEBUG] header_row_index = {header_row_index}")
    debug_print(f"[DEBUG] df.columns = {list(df.columns)}")

    detected = {}
    normalized_cols = [normalize(c) for c in df.columns]
    debug_print(f"[DEBUG] normalized_cols = {normalized_cols}")

    for target in target_columns:
        norm_target = normalize(target)
        debug_print(f"[DEBUG] Searching for '{target}' (normalized='{norm_target}')")
        for i, col in enumerate(normalized_cols):
            if norm_target in col or col in norm_target:
                debug_print(f"[DEBUG]  -> matched '{df.columns[i]}' (normalized='{col}')")
                detected[target] = i
                break
            if norm_target.startswith("pg") and ("pg" in col or "ｐｇ" in col):
                debug_print(f"[DEBUG]  -> matched (special rule) '{df.columns[i]}' (normalized='{col}')")
                detected[target] = i
                break
    debug_print(f"[DEBUG] detect_columns detected={detected}")
    debug_print(f"[DEBUG] ===== detect_columns end =====\n")

    return detected


# ========= 数量セルの背景色を取得 =========
def get_quantity_colors(file_path: str, sheet_name: str, header_row_index: int, quantity_col_index: int):
    """
    数量セルに「何らかの色が付いているかどうか」を判定して返す。
    - パターンなし or 塗りなし → ""（無色扱い）
    - 何らかの色（RGB / indexed / theme） → 何かしらの文字列（非空）
    """
    wb = load_workbook(file_path, data_only=True)

    try:
        ws = wb[sheet_name]
        colors = []

        for row_idx in range(header_row_index + 2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=quantity_col_index + 1)  # openpyxlは1始まり
            fill = cell.fill

            if not fill:
                colors.append("")
                continue

            pattern = getattr(fill, "patternType", None)

            # パターンが none / 未設定なら「塗りなし」とみなす
            if pattern in (None, "none"):
                colors.append("")
                continue

            fg = fill.fgColor
            if fg is None:
                colors.append("")
                continue

            colored = False
            marker = ""

            # 1) RGB 色
            if fg.type == "rgb":
                # 完全な白 or 透明っぽい値は「塗りなし」とみなす
                if fg.rgb not in (None, "00000000", "FFFFFFFF"):
                    colored = True
                    marker = fg.rgb

            # 2) インデックス色（パレット）
            elif fg.type == "indexed":
                # 0, 64 は「塗りなし」的な扱いが多いので除外
                if fg.index not in (0, 64):
                    colored = True
                    marker = f"indexed:{fg.index}"

            # 3) テーマ色
            elif fg.type == "theme":
                # テーマ色で塗りがある場合は、とりあえず「色あり」とみなす
                colored = True
                marker = f"theme:{fg.theme}"

            # その他タイプは一応「色なし」として扱う
            if colored:
                colors.append(marker)
            else:
                colors.append("")

        return colors

    finally:
        wb.close()


# ========= Excel抽出処理 =========
def extract_data_from_excel(file_path: str, sheet_name: str, detected_columns: dict, header_row_index: int):
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row_index)

    col_items = []
    for key in ["品番", "PG名", "品名", "材料", "数量"]:
        if key in detected_columns:
            col_name = df.columns[detected_columns[key]]
            col_items.append(col_name)

    sub_df = df[col_items].copy() if col_items else pd.DataFrame()

    # --- 品名列が結合されている場合への対応 ---
    cols = list(df.columns)

    # --- 共通列への対応 ---
    # 「共通」は内部処理には使わず、検索結果テーブルに表示するだけの項目。
    # Excel上に「共通」を含む列があれば取得する。
    common_col = next((c for c in df.columns if "共通" in str(c)), None)

    if common_col:
        sub_df["共通"] = (
            df[common_col]
            .fillna("")
            .astype(str)
            .str.strip()
            .replace({"nan": "", "None": "", "none": ""})
        )

    def find_name_block_indices(cols):
        for i, c in enumerate(cols):
            if "品名" in str(c):
                j = i + 1
                block = [i]
                while j < len(cols):
                    cj = str(cols[j])
                    if cj.startswith("Unnamed") or cj.strip() == "" or cj.lower() in ("none", "nan"):
                        block.append(j)
                        j += 1
                    else:
                        break
                return block
        return []

    name_idx = find_name_block_indices(cols)
    if name_idx:
        def clean_name_value(value) -> str:
            if pd.isna(value):
                return ""

            s = str(value).strip()

            if s == "" or s.lower() in ("nan", "none"):
                return ""

            return s

        def join_name_parts(parts: list[str]) -> str:
            return " ".join(p for p in parts if p).strip()

        # 品名列が1列だけの場合は、今まで通りその列の値を品名にする
        if len(name_idx) == 1:
            first_col = df.columns[name_idx[0]]
            sub_df["品名"] = df[first_col].apply(clean_name_value)

        # 品名列が2列以上ある場合は、
        # 1列目を「大品名」としてWK保持し、2列目以降のみの行に引き継ぐ
        else:
            first_name_col = df.columns[name_idx[0]]
            detail_name_cols = [df.columns[i] for i in name_idx[1:]]

            current_major_name = ""
            item_names = []

            for idx, row in df.iterrows():
                major_name = clean_name_value(row.get(first_name_col, ""))
                detail_parts = [
                    clean_name_value(row.get(col, ""))
                    for col in detail_name_cols
                ]
                detail_name = join_name_parts(detail_parts)

                # 品名1列目に値があれば、数量の有無に関係なくWK大品名を更新する
                if major_name:
                    current_major_name = major_name

                if major_name and detail_name:
                    # 同じ行に1列目・2列目両方がある場合は2段表示にする
                    item_name = "\n".join([major_name, detail_name])

                elif major_name:
                    # 1列目のみの通常行
                    item_name = major_name

                elif detail_name:
                    # 2列目以降のみの行は、直近の大品名を前に付けて2段表示にする
                    if current_major_name:
                        item_name = "\n".join([current_major_name, detail_name])
                    else:
                        item_name = detail_name

                else:
                    item_name = ""

                item_names.append(item_name)

            sub_df["品名"] = item_names

    else:
        name_like_cols = [c for c in df.columns if "品名" in str(c)]
        if len(name_like_cols) == 1:
            sub_df["品名"] = (
                df[name_like_cols[0]]
                .apply(lambda v: "" if pd.isna(v) else str(v).strip())
                .replace({"nan": "", "None": "", "none": ""})
            )

    # --- 数量フィルタ ---
    # detect_columns() で検出した数量列を使う。
    # 数量欄が完全に空（NaN / 空白）の行だけ除外し、0 は残す。
    quantity_col = None
    if "数量" in detected_columns:
        quantity_col = df.columns[detected_columns["数量"]]

    # --- PG名のみの行への前行情報引き継ぎ ---
    # Excel上で品名・材料・品番・数量・共通などが縦結合され、
    # 2行目にPG名だけが入っているケースに対応する。
    #
    # 数量が空白でもPG名がある行は、直前行の情報を引き継いで
    # 抽出対象として残す。
    def is_blank_cell_value(value) -> bool:
        if pd.isna(value):
            return True

        s = str(value).strip()

        return s == "" or s.lower() in ("nan", "none")

    pg_col = None
    if "PG名" in detected_columns:
        pg_col = df.columns[detected_columns["PG名"]]

    def fill_from_previous_row_if_blank(target_col: str, idx, prev_idx):
        """
        target_col が sub_df に存在し、現在行が空白なら前行の値を引き継ぐ。
        """
        if target_col not in sub_df.columns:
            return

        current_value = sub_df.at[idx, target_col]
        previous_value = sub_df.at[prev_idx, target_col]

        if is_blank_cell_value(current_value) and not is_blank_cell_value(previous_value):
            sub_df.at[idx, target_col] = previous_value

    if quantity_col is not None and pg_col is not None:
        for pos, idx in enumerate(sub_df.index):
            if pos == 0:
                continue

            prev_idx = sub_df.index[pos - 1]

            quantity_blank = is_blank_cell_value(df.at[idx, quantity_col])
            pg_exists = not is_blank_cell_value(df.at[idx, pg_col])

            if quantity_blank and pg_exists:
                # 品名・材料・品番・共通は、現在行が空白なら前行から補完する
                for target_col in ["品名", "材料", "品番", "共通"]:
                    fill_from_previous_row_if_blank(target_col, idx, prev_idx)

                # 数量は、この行を抽出対象として残すために前行から補完する
                if "数量" in sub_df.columns:
                    previous_quantity = sub_df.at[prev_idx, "数量"]

                    if not is_blank_cell_value(previous_quantity):
                        sub_df.at[idx, "数量"] = previous_quantity

    if quantity_col is not None:
        def keep_row_by_quantity(value):
            if pd.isna(value):
                return False

            s = str(value).strip()

            if s == "" or s.lower() in ("nan", "none"):
                return False

            return True

        # 前行補完後の sub_df["数量"] を見てフィルタする。
        # これにより、Excel上の数量欄が空白でも、PG名ありで前行数量を引き継いだ行は残る。
        if "数量" in sub_df.columns:
            sub_df = sub_df[sub_df["数量"].apply(keep_row_by_quantity)]
        else:
            sub_df = sub_df[df[quantity_col].apply(keep_row_by_quantity)]

    # ✅ 検出された列に基づいて、列名を統一（英字ブレ対応）
    rename_map = {}
    if "品番" in detected_columns:
        rename_map[df.columns[detected_columns["品番"]]] = "品番"
    if "PG名" in detected_columns:
        rename_map[df.columns[detected_columns["PG名"]]] = "PG名"
    if "品名" in detected_columns:
        rename_map[df.columns[detected_columns["品名"]]] = "品名"
    if "材料" in detected_columns:
        rename_map[df.columns[detected_columns["材料"]]] = "材料"
    if "数量" in detected_columns:
        rename_map[df.columns[detected_columns["数量"]]] = "数量"

    sub_df.rename(columns=rename_map, inplace=True)

    # ✅ 品名列が複数（例：'品名', 'Unnamed:3'）ある場合は結合して1列に統一
    name_like_cols = [c for c in sub_df.columns if "品名" in str(c) or "Unnamed" in str(c)]
    if len(name_like_cols) > 1:
        sub_df["品名"] = sub_df[name_like_cols].astype(str).apply(
            lambda row: " ".join(v.strip() for v in row if v.strip() and v.lower() != "nan"),
            axis=1
        )
        sub_df.drop(columns=[c for c in name_like_cols if c != "品名"], inplace=True)

    # ✅ 対象列のみを保持（項番・備考など除外）
    keep_cols = [c for c in ["品番", "PG名", "品名", "材料", "共通", "数量"] if c in sub_df.columns]
    sub_df = sub_df[keep_cols]

    return sub_df

# ========= メイン =========
def main(page: ft.Page):
    page.title = f"図面PDF検索結合ツール {APP_VERSION}"
    page.theme_mode = ft.ThemeMode.DARK

    icon_path = resource_path(APP_ICON_FILE)
    if icon_path.is_file():
        page.window.icon = str(icon_path.resolve())

    # 起動時のウィンドウ表示設定
    try:
        page.window.maximized = True
        page.window.width = 1400
        page.window.height = 900
        page.window.min_width = 1200
        page.window.min_height = 700
    except Exception:
        pass

    # プレビューキャッシュを起動時に削除
    clear_preview_cache()

    sheet_dropdown = None

    # DataTable状態管理用
    current_df = pd.DataFrame()
    search_text_fields = {}
    output_checkboxes = {}

    HOVER_PREVIEW_MIN_WIDTH = 1700
    use_hover_preview = False

    # ---- 設定ファイル処理 ----
    def load_config():
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        return {
            "pdf_folder": "",
            "output_folder": "",
            "excel_folder": "",
            "search_mode": "構成部品表",
            "target_columns": ["品番", "PG名"],
        }

    def save_config(e=None):
        """
        現在の設定を config.json に保存する。

        v1.0.2 以降は、設定変更時に自動保存するため、
        保存完了メッセージは表示しない。
        """
        config.update({
            "pdf_folder": search_folder_field.value,
            "output_folder": output_folder_field.value,
            "excel_folder": config.get("excel_folder", ""),
            "search_mode": mode_dropdown.value,
            "target_columns": [x.strip() for x in target_col_field.value.split(",") if x.strip()],
        })

        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(config, f, indent=2, ensure_ascii=False)

    def close_app(e):
        page.window.close()

    def create_button_style(bgcolor, color=ft.Colors.WHITE):
        """
        ダークテーマ上で見やすいボタンスタイルを作る。
        """
        return ft.ButtonStyle(
            bgcolor=bgcolor,
            color=color,
            padding=ft.padding.symmetric(horizontal=10, vertical=6),
        )


    BUTTON_STYLE_PRIMARY = create_button_style(ft.Colors.BLUE_700)
    BUTTON_STYLE_SUCCESS = create_button_style(ft.Colors.GREEN_700)
    BUTTON_STYLE_WARNING = create_button_style(ft.Colors.ORANGE_700)
    BUTTON_STYLE_SECONDARY = create_button_style(ft.Colors.BLUE_GREY_700)
    BUTTON_STYLE_DANGER_OUTLINE = ft.ButtonStyle(
        color=ft.Colors.RED_100,
        side=ft.BorderSide(1, ft.Colors.RED_300),
        padding=ft.padding.symmetric(horizontal=12, vertical=8),
    )

    config = load_config()

    # ---- UI構成 ----
    search_folder_field = ft.TextField(label="検索先フォルダ", value=config.get("pdf_folder", ""), expand=True)
    output_folder_field = ft.TextField(label="出力先フォルダ", value=config.get("output_folder", ""), expand=True)

    search_folder_field.on_change = save_config
    output_folder_field.on_change = save_config

    target_col_field = ft.TextField(
        label="抽出対象列（カンマ区切り）",
        value=", ".join(config.get("target_columns", [])),
        width=450  # ← expand=False & 半分幅
    )
    mode_dropdown = ft.Dropdown(
        label="検索モード",
        options=[ft.dropdown.Option("構成部品表"), ft.dropdown.Option("通常")],
        value=config.get("search_mode", "構成部品表"),
        width=150
    )

    exit_button = ft.OutlinedButton(
        "終了",
        icon=ft.Icons.CLOSE,
        on_click=close_app,
        style=BUTTON_STYLE_DANGER_OUTLINE,
    )

    settings_action_row = ft.Row(
        controls=[
            exit_button,
        ],
        spacing=8,
    )

    # ---- フォルダ選択ダイアログ ----
    folder_picker_search = ft.FilePicker(on_result=lambda e: pick_search_result(e))
    folder_picker_output = ft.FilePicker(on_result=lambda e: pick_output_result(e))
    page.overlay.extend([folder_picker_search, folder_picker_output])

    def pick_search_folder(e):
        start_path = search_folder_field.value.strip()  # ← テキストボックスから取得
        if not os.path.isdir(start_path):
            start_path = os.getcwd()
        folder_picker_search.get_directory_path(initial_directory=start_path)

    def pick_output_folder(e):
        start_path = output_folder_field.value.strip()  # ← テキストボックスから取得
        if not os.path.isdir(start_path):
            start_path = os.getcwd()
        folder_picker_output.get_directory_path(initial_directory=start_path)

    def pick_search_result(e: ft.FilePickerResultEvent):
        if e.path:
            search_folder_field.value = e.path
            save_config()
            page.update()

    def pick_output_result(e: ft.FilePickerResultEvent):
        if e.path:
            output_folder_field.value = e.path
            save_config()
            page.update()

    # ---- 検索モード行 ----
    def update_mode_fields():
        if mode_dropdown.value == "構成部品表":
            mode_row.controls = [
                ft.Row(
                    controls=[
                        mode_dropdown,
                    ],
                    spacing=12,
                ),
                settings_action_row,
            ]
            set_excel_ui_enabled(True)

        else:
            manual_field = ft.TextField(
                label="検索文字列（手動入力）",
                width=450,
            )
            mode_row.controls = [
                ft.Row(
                    controls=[
                        mode_dropdown,
                        manual_field,
                    ],
                    spacing=12,
                ),
                settings_action_row,
            ]
            set_excel_ui_enabled(False)

        page.update()

    def on_mode_change(e):
        save_config()
        update_mode_fields()

    mode_row = ft.Row(
        controls=[],
        alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
    )
    mode_dropdown.on_change = on_mode_change

    # ---- Excel選択 ----
    selected_excel_path = ""
    SHEET_PLACEHOLDER = "（シートを選択してください）"
    sheet_dropdown = ft.Dropdown(label="シート選択", width=420)
    message = ft.Text(
        "",
        size=16,
        weight=ft.FontWeight.BOLD,
        color=ft.Colors.BLUE_100,
    )
    
    mode_notice = ft.Text("")
    table = ft.DataTable(
        columns=[ft.DataColumn(ft.Text("抽出結果"))],
        rows=[],
        column_spacing=20,
        horizontal_margin=12,
        data_row_min_height=46,
        data_row_max_height=72,
    )
    table_header = ft.Row([], alignment="center")

    hover_preview_image = ft.Image(
        src="",
        width=820,
        height=580,
        fit=ft.ImageFit.CONTAIN,
        visible=False,
    )

    hover_preview_message = ft.Text(
        "",
        size=12,
        color=ft.Colors.BLUE_700,
        weight=ft.FontWeight.BOLD,
        text_align=ft.TextAlign.CENTER,
    )

    hover_preview_panel = ft.Container(
        padding=ft.padding.only(left=2, right=2, top=2),
        visible=False,
        bgcolor=ft.Colors.TRANSPARENT,
        expand=True,
    )

    def format_quantity_for_display(v) -> str:
        if pd.isna(v):
            return ""
        s = str(v).strip()
        if s == "" or s.lower() in ("nan", "none"):
            return ""

        # "1.0" のような整数の小数表現だけ "1" に戻す
        try:
            x = float(s.replace(",", ""))
            if x.is_integer():
                return str(int(x))
        except ValueError:
            pass

        return s

    def is_zero_quantity(v) -> bool:
        if pd.isna(v):
            return False

        s = str(v).strip()

        if s == "":
            return False

        try:
            return float(s.replace(",", "")) == 0.0
        except ValueError:
            return False

    def is_output_ineligible(row) -> bool:
        """
        出力可否が「不可」の行かどうかを返す。
        """
        return str(row.get("出力可否", "") or "").strip() == "不可"

    def create_preview_box(preview_image: ft.Image, height: int = 520) -> ft.Container:
        """
        PDFプレビュー画像を表示するための共通枠を作る。
        """
        return ft.Container(
            content=preview_image,
            alignment=ft.alignment.center,
            border=ft.border.all(1, ft.Colors.GREY_500),
            border_radius=8,
            padding=10,
            height=height,
            bgcolor=ft.Colors.GREY_200,
        )

    def create_hover_preview_box(preview_image: ft.Image, height: int = 620) -> ft.Container:
        return ft.Container(
            content=preview_image,
            alignment=ft.alignment.top_center,
            height=height,
            padding=0,
            bgcolor=ft.Colors.TRANSPARENT,
        )

    hover_preview_panel.content = ft.Container(
        content=ft.Column(
            controls=[
                hover_preview_message,
                create_hover_preview_box(hover_preview_image, height=600)
            ],
            spacing=0,
            tight=True,
            horizontal_alignment=ft.CrossAxisAlignment.CENTER,
        ),
        alignment=ft.alignment.top_center,
        expand=True,
    )

    def update_hover_preview(pdf_path: str):
        """
        出力PDF欄にホバーしたとき、右側の固定プレビュー欄を更新する。
        """
        if not use_hover_preview:
            return

        if not pdf_path or not str(pdf_path).strip():
            return

        hover_preview_image.visible = False
        hover_preview_message.value = "プレビューを読み込んでいます..."
        page.update()

        try:
            preview_path = create_pdf_preview_image(
                pdf_path,
                page_number=0,
                dpi=100,
            )

            hover_preview_image.src = preview_path
            hover_preview_image.visible = True
            hover_preview_message.value = Path(pdf_path).name

        except Exception:
            hover_preview_image.src = ""
            hover_preview_image.visible = False
            hover_preview_message.value = "プレビューを生成できませんでした。"

        page.update()

    def clear_hover_preview():
        """
        ホバープレビュー画像とファイル名表示をクリアする。
        """
        hover_preview_image.src = ""
        hover_preview_image.visible = False
        hover_preview_message.value = ""

    def on_preview_adopted_pdf(pdf_path: str):
        """
        採用済みPDFをプレビュー表示する。
        自動採用・手動採用どちらでも、採用PDFパスがある行から呼び出す。
        """
        if not pdf_path or not str(pdf_path).strip():
            page.open(ft.SnackBar(ft.Text("採用PDFがありません。")))
            page.update()
            return

        preview_image = ft.Image(
            src="",
            width=760,
            height=520,
            fit=ft.ImageFit.CONTAIN,
            visible=False,
        )

        preview_message = ft.Text("プレビューを読み込み中...")

        def load_preview():
            preview_image.visible = False
            preview_message.value = "プレビューを読み込んでいます..."
            page.update()

            try:
                preview_path = create_pdf_preview_image(
                    pdf_path,
                    page_number=0,
                    dpi=120,
                )
                preview_image.src = preview_path
                preview_image.visible = True
                preview_message.value = format_pdf_path_for_display(pdf_path)

            except Exception as ex:
                preview_image.src = ""
                preview_image.visible = False
                preview_message.value = (
                    "プレビューを生成できませんでした。"
                    f" PDFにアクセスできない、または形式に問題がある可能性があります。詳細: {ex}"
                )

            page.update()

        def close_dialog(e):
            render_table_from_df(current_df)
            dialog.open = False
            page.update()

        dialog = ft.AlertDialog(
            modal=True,
            title=ft.Text("採用PDFプレビュー"),
            content=ft.Container(
                width=860,
                height=640,
                content=ft.Column(
                    controls=[
                        preview_message,
                        create_preview_box(preview_image, height=560),
                    ],
                    tight=True,
                ),
            ),
            actions=[
                ft.TextButton("再読み込み", on_click=lambda e: load_preview()),
                ft.TextButton("閉じる", on_click=close_dialog),
            ],
            actions_alignment="end",
        )

        page.open(dialog)
        page.update()

        load_preview()

    def on_select_pdf_candidate(row_index):
        """
        複数候補のPDFから、採用するPDFを1つ選択する。
        ラジオボタンで選択した内容を、即時に採用PDFパスへ反映する。
        """
        nonlocal current_df

        if current_df is None or current_df.empty:
            page.open(ft.SnackBar(ft.Text("抽出結果がありません。")))
            page.update()
            return

        sync_table_state_to_current_df()

        if row_index not in current_df.index:
            page.open(ft.SnackBar(ft.Text("対象行が見つかりません。")))
            page.update()
            return

        row = current_df.loc[row_index]

        def format_dialog_value(value) -> str:
            if pd.isna(value) or str(value).strip().lower() in ("nan", "none"):
                return ""
            return str(value).strip()

        candidates = row.get("候補PDFパス一覧", [])

        if not isinstance(candidates, list):
            candidates = []

        if not candidates:
            page.open(ft.SnackBar(ft.Text("候補PDFはありません。")))
            page.update()
            return

        current_adopted_pdf = str(row.get("採用PDFパス", "") or "").strip()

        preview_image = ft.Image(
            src="",
            width=980,
            height=660,
            fit=ft.ImageFit.CONTAIN,
            visible=False,
        )

        preview_message = ft.Text("左の候補PDFを選択すると、ここにプレビューを表示します。")

        def update_preview(pdf_path: str):
            if not pdf_path:
                preview_image.src = ""
                preview_image.visible = False
                preview_message.value = "左の候補PDFを選択すると、ここにプレビューを表示します。"
                page.update()
                return

            preview_image.visible = False
            preview_message.value = "プレビューを読み込んでいます..."
            page.update()

            try:
                preview_path = create_pdf_preview_image(
                    pdf_path,
                    page_number=0,
                    dpi=120,
                )
                preview_image.src = preview_path
                preview_image.visible = True
                preview_message.value = format_pdf_path_for_display(pdf_path)

            except Exception as ex:
                preview_image.src = ""
                preview_image.visible = False
                preview_message.value = (
                    "プレビューを生成できませんでした。"
                    f" PDFにアクセスできない、または形式に問題がある可能性があります。詳細: {ex}"
                )

            page.update()

        NO_ADOPTED_PDF_VALUE = "__NO_ADOPTED_PDF__"

        def apply_candidate_selection(selected_value: str):
            """
            ラジオボタンで選択された候補を、即時に採用状態へ反映する。
            """
            sync_table_state_to_current_df()

            if selected_value == NO_ADOPTED_PDF_VALUE:
                current_df.at[row_index, "採用PDFパス"] = ""
                current_df.at[row_index, "候補状態"] = "複数候補"

                preview_image.src = ""
                preview_image.visible = False
                preview_message.value = "採用しない状態です。左の候補PDFを選ぶとプレビューを表示します。"

                debug_print("===== PDF候補 採用解除 =====")
                debug_print(f"行index: {row_index}")
                debug_print(f"検索文字列: {row.get('検索用文字列', '')!r}")

                page.update()
                return

            if not selected_value:
                return

            current_df.at[row_index, "採用PDFパス"] = selected_value
            current_df.at[row_index, "候補状態"] = "手動採用"

            debug_print("===== PDF候補 即時採用 =====")
            debug_print(f"行index: {row_index}")
            debug_print(f"検索文字列: {row.get('検索用文字列', '')!r}")
            debug_print(f"採用PDF: {format_pdf_path_for_display(selected_value)}")
            debug_print(f"採用PDFフルパス: {selected_value}")

            update_preview(selected_value)

        def create_candidate_pdf_option(pdf_path: str) -> ft.Container:
            """
            候補選択ダイアログ用のPDF候補表示を作る。

            1行目: フォルダパス
            2行目: ファイル名

            ホバー時:
                右側プレビューを更新する。

            クリック時:
                ラジオボタンをONにして採用PDFへ反映する。
            """
            folder_text, file_name = split_pdf_path_for_two_line_display(pdf_path)

            radio = ft.Radio(value=pdf_path)

            file_name_text = ft.Text(
                file_name,
                color=PDF_FILE_NAME_COLOR,
                size=13,
                weight=ft.FontWeight.BOLD,
                no_wrap=True,
            )

            text_column = ft.Column(
                controls=[
                    ft.Text(
                        folder_text,
                        color=ft.Colors.WHITE,
                        size=12,
                        no_wrap=True,
                    ),
                    file_name_text,
                ],
                spacing=0,
                tight=True,
            )

            def preview_candidate_on_hover(e, preview_pdf_path=pdf_path):
                """
                候補PDFにマウスを乗せたとき、右側プレビューを更新し、
                ファイル名の色を少し変える。
                """
                if e.data == "true":
                    file_name_text.color = PDF_FILE_NAME_HOVER_COLOR
                    update_preview(preview_pdf_path)
                else:
                    file_name_text.color = PDF_FILE_NAME_COLOR

                page.update()

            def select_candidate(e, selected_value=pdf_path):
                selected_pdf.value = selected_value
                apply_candidate_selection(selected_value)
                page.update()

            return ft.Container(
                content=ft.Row(
                    controls=[
                        radio,
                        text_column,
                    ],
                    spacing=6,
                    vertical_alignment=ft.CrossAxisAlignment.CENTER,
                ),
                padding=ft.padding.symmetric(horizontal=4, vertical=3),
                border_radius=6,
                on_hover=preview_candidate_on_hover,
                on_click=select_candidate,
            )

        candidate_list_height = min(
            460,
            max(60, len(candidates) * 48),
        )

        selected_pdf = ft.RadioGroup(
            value=current_adopted_pdf if current_adopted_pdf in candidates else NO_ADOPTED_PDF_VALUE,
            on_change=lambda e: apply_candidate_selection(e.control.value),
            content=ft.Column(
                controls=[
                    ft.Container(
                        height=candidate_list_height,
                        content=ft.Column(
                            controls=[
                                create_candidate_pdf_option(pdf_path)
                                for pdf_path in candidates
                            ],
                            spacing=2,
                            tight=True,
                            scroll="auto",
                        ),
                    ),
                    ft.Divider(height=8),
                    ft.Container(
                        content=ft.Radio(
                            value=NO_ADOPTED_PDF_VALUE,
                            label="採用しない",
                        ),
                        padding=ft.padding.symmetric(horizontal=4, vertical=3),
                        border_radius=6,
                        on_click=lambda e: (
                            setattr(selected_pdf, "value", NO_ADOPTED_PDF_VALUE),
                            apply_candidate_selection(NO_ADOPTED_PDF_VALUE),
                            page.update(),
                        ),
                    ),
                ],
                spacing=2,
                tight=True,
            ),
        )

        def close_dialog(e):
            nonlocal current_df

            current_df = apply_output_target_defaults(current_df)
            current_df = update_adopted_pdf_duplicate_info(current_df)
            current_df = apply_duplicate_output_defaults(current_df)

            render_table_from_df(current_df)
            dialog.open = False
            page.update()

        ok_button = ft.ElevatedButton(
            "OK",
            on_click=close_dialog,
        )

        dialog = ft.AlertDialog(
            modal=True,
            title=ft.Text("PDF候補を選択"),
            content=ft.Container(
                width=1740,
                height=800,
                content=ft.Row(
                    controls=[
                        ft.Container(
                            width=500,
                            content=ft.Column(
                                controls=[
                                    ft.Text(
                                        f"検索文字列: {format_dialog_value(row.get('検索用文字列', ''))}",
                                        weight="bold",
                                    ),
                                    ft.Text(f"品名: {format_dialog_value(row.get('品名', ''))}"),
                                    ft.Text(f"材料: {format_dialog_value(row.get('材料', ''))}"),
                                    ft.Text(f"候補数: {len(candidates)}"),
                                    ft.Text(f"現在の状態: {format_dialog_value(row.get('候補状態', ''))}"),
                                    ft.Divider(),
                                    ft.Text("採用するPDFを選んでください。選択はすぐに反映されます。"),
                                    selected_pdf,
                                    ft.Row(
                                        controls=[ok_button],
                                        alignment=ft.MainAxisAlignment.END,
                                    ),
                                ],
                                spacing=4,
                                tight=True,
                            ),
                        ),
                        ft.VerticalDivider(),
                        ft.Container(
                            width=1060,
                            content=ft.Column(
                                controls=[
                                    ft.Text("プレビュー", weight="bold"),
                                    preview_message,
                                    create_preview_box(preview_image, height=700),
                                ],
                                tight=True,
                            ),
                        ),
                    ],
                    vertical_alignment=ft.CrossAxisAlignment.START,
                    scroll="auto",
                ),
            ),
        )

        initial_preview_pdf = selected_pdf.value

        page.open(dialog)
        page.update()

        if initial_preview_pdf and initial_preview_pdf != NO_ADOPTED_PDF_VALUE:
            update_preview(initial_preview_pdf)

    def update_adopted_pdf_duplicate_info(df: pd.DataFrame) -> pd.DataFrame:
        """
        採用PDFパスの重複状況を表示用列に反映する。

        同じ採用PDFパスが複数行に入っている場合:
        - 1件目: 「1件目」
        - 2件目以降: 「2件目→」「3件目→」...

        採用PDFパスが空欄の行は重複判定しない。
        """
        if df is None or df.empty:
            return df

        if "採用PDFパス" not in df.columns:
            return df

        df = df.copy()

        adopted_paths = df["採用PDFパス"].fillna("").astype(str).str.strip()

        # 初期値は空欄
        df["採用PDF重複"] = ""

        # 空欄は重複判定から除外
        non_empty = adopted_paths.ne("")

        if not non_empty.any():
            return df

        counts = adopted_paths[non_empty].value_counts()
        duplicate_paths = counts[counts >= 2]

        if duplicate_paths.empty:
            return df

        for adopted_path, count in duplicate_paths.items():
            matching_indices = df.index[non_empty & adopted_paths.eq(adopted_path)].tolist()

            for position, idx in enumerate(matching_indices, start=1):
                if position == 1:
                    df.at[idx, "採用PDF重複"] = "1件目"
                else:
                    df.at[idx, "採用PDF重複"] = f"{position}件目→"

        return df

    def apply_duplicate_output_defaults(df: pd.DataFrame) -> pd.DataFrame:
        """
        採用PDFパスが重複している場合、2件目以降を出力対象外にする。

        - 重複なし: 既存の出力対象を維持
        - 1件目: 出力対象を維持
        - 2件目以降: 出力対象を False にする
        """
        if df is None or df.empty:
            return df

        if "採用PDF重複" not in df.columns or "出力対象" not in df.columns:
            return df

        df = df.copy()

        duplicate_info = df["採用PDF重複"].fillna("").astype(str).str.strip()

        # 「2件目→」「3件目→」など、矢印付きの行をOFFにする
        second_or_later = duplicate_info.str.endswith("→")

        df.loc[second_or_later, "出力対象"] = False

        return df

    def format_pdf_path_for_display(pdf_path: str) -> str:
        """
        PDFパスを画面表示用に整形する。
        検索先フォルダ配下のPDFであれば、検索先フォルダ以降の相対パスで表示する。
        """
        if not pdf_path or not str(pdf_path).strip():
            return ""

        path_text = str(pdf_path).strip()
        search_root = search_folder_field.value.strip()

        if not search_root:
            return path_text

        try:
            pdf_path_obj = Path(path_text)
            search_root_obj = Path(search_root)

            relative_path = pdf_path_obj.relative_to(search_root_obj)
            return str(relative_path)

        except ValueError:
            return path_text
        except Exception:
            return path_text

    def split_pdf_path_for_two_line_display(pdf_path: str) -> tuple[str, str]:
        """
        出力PDF欄を2段表示するために、PDFパスをフォルダ部分とファイル名に分ける。

        1行目: フォルダパス
        2行目: PDFファイル名
        """
        display_path = format_pdf_path_for_display(pdf_path)

        if not display_path:
            return "", ""

        path_obj = Path(display_path)

        folder_text = str(path_obj.parent)

        if folder_text == ".":
            folder_text = ""

        file_name = path_obj.name

        return folder_text, file_name

    def render_table_from_df(df: pd.DataFrame):
        nonlocal current_df

        current_df = df.copy()

        search_text_fields.clear()
        output_checkboxes.clear()

        display_df = df.copy()
        display_df.insert(0, "項番", range(1, len(display_df) + 1))

        update_hover_preview_visibility()

        hidden_columns = {
            "候補PDFパス一覧",
            "先頭候補PDF",
            "元検索文字列",
            "候補状態",
            "候補PDF数",
            "縦結合",
            "品番",
            "PG名",
        }

        preferred_order = [
            "項番",
            "品名",
            "材料",
            "品番PG名",
            "数量",
            "数量セル色",
            "出力可否",
            "検索用文字列",
            "共通",
            "候補確認",
            "採用PDFパス",
            "採用PDF重複",
            "出力対象",
            "プレビュー",
        ]

        display_only_columns = {
            "品番PG名",
            "候補確認",
            "プレビュー",
        }

        display_columns = [
            c for c in preferred_order
            if (c in display_df.columns or c in display_only_columns) and c not in hidden_columns
        ]

        display_columns += [
            c for c in display_df.columns
            if c not in display_columns and c not in hidden_columns
        ]

        # 余白ホバープレビューが使える画面では、クリック用プレビュー列は表示しない
        if use_hover_preview:
            display_columns = [
                c for c in display_columns
                if c != "プレビュー"
            ]

        def get_display_column_name(column_name: str) -> str:
            display_name_map = {
                "検索用文字列": "検索文字列",
                "採用PDF重複": "重複",
                "数量": "数量ゼロ",
                "数量セル色": "セル色",
                "採用PDFパス": "出力PDF",
                "出力可否": "抽出対象",
                "出力対象": "出力",
                "候補確認": "複数候補",
                "品番PG名": "品番/PG名",
            }
            return display_name_map.get(column_name, column_name)

        def get_header_badge_counts(df: pd.DataFrame) -> dict:
            """
            テーブル見出し横に表示する件数を返す。
            キーは内部列名。
            """
            counts = {}

            if df is None or df.empty:
                return counts

            if "出力可否" in df.columns:
                counts["出力可否"] = int(
                    df["出力可否"]
                    .fillna("")
                    .astype(str)
                    .str.strip()
                    .eq("可")
                    .sum()
                )

            if "候補PDF数" in df.columns:
                candidate_counts = pd.to_numeric(
                    df["候補PDF数"],
                    errors="coerce",
                ).fillna(0)

                # 候補確認が必要な件数として、複数候補だけ数える
                counts["候補確認"] = int((candidate_counts >= 2).sum())

            if "出力対象" in df.columns:
                counts["出力対象"] = int(
                    df["出力対象"]
                    .fillna(False)
                    .astype(bool)
                    .sum()
                )

            return counts

        def get_header_badge_color(column_name: str):
            color_map = {
                "出力可否": ft.Colors.BLUE_100,
                "候補確認": ft.Colors.ORANGE_100,
                "出力対象": ft.Colors.GREEN_100,
            }
            return color_map.get(column_name, ft.Colors.GREY_300)

        def create_header_label(column_name: str, badge_counts: dict):
            display_name = get_display_column_name(column_name)
            count = badge_counts.get(column_name)

            compact_header_lines = {
                "数量": ["数量", "ゼロ"],
                "数量セル色": ["セル", "色"],
                "出力可否": ["抽出", "対象"],
                "候補確認": ["複数", "候補"],
            }

            def create_header_text():
                lines = compact_header_lines.get(column_name)

                if lines:
                    return ft.Column(
                        controls=[
                            ft.Text(
                                line,
                                size=12,
                                weight=ft.FontWeight.BOLD,
                                text_align=ft.TextAlign.CENTER,
                            )
                            for line in lines
                        ],
                        spacing=0,
                        tight=True,
                        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                    )

                return ft.Text(
                    display_name,
                    size=12,
                    weight=ft.FontWeight.BOLD,
                    text_align=ft.TextAlign.CENTER,
                )

            header_text = create_header_text()

            # 検索文字列列は、入力欄の幅に合わせて見出しを中央寄せする
            if column_name == "検索用文字列":
                return ft.Container(
                    width=100,
                    alignment=ft.alignment.center,
                    content=header_text,
                )

            if count is None:
                return header_text

            badge = ft.Container(
                content=ft.Text(
                    str(count),
                    size=12,
                    color=ft.Colors.BLACK,
                    weight=ft.FontWeight.BOLD,
                ),
                padding=ft.padding.symmetric(horizontal=7, vertical=3),
                border_radius=12,
                bgcolor=get_header_badge_color(column_name),
            )

            return ft.Row(
                controls=[
                    header_text,
                    badge,
                ],
                spacing=4,
                alignment=ft.MainAxisAlignment.CENTER,
                vertical_alignment=ft.CrossAxisAlignment.CENTER,
            )

        def should_center_align_column(column_name: str) -> bool:
            center_columns = {
                "項番",
                "共通",
                "数量",
                "数量セル色",
                "出力可否",
                "採用PDF重複",
                "出力対象",
                "候補PDF数",
                "候補確認",
                "プレビュー",
            }
            return column_name in center_columns

        badge_counts = get_header_badge_counts(current_df)

        table.columns = [
            ft.DataColumn(create_header_label(c, badge_counts))
            for c in display_columns
        ]
        table.rows = []

        ineligible_blank_columns = {
            "検索用文字列",
            "共通",
            "採用PDF重複",
            "出力対象",
            "候補PDF数",
            "候補確認",
            "採用PDFパス",
            "プレビュー",
        }

        def on_output_checkbox_change(e, row_index):
            nonlocal current_df

            if current_df is None or current_df.empty:
                return

            if row_index in current_df.index:
                current_df.at[row_index, "出力対象"] = bool(e.control.value)

            render_table_from_df(current_df)
            page.update()

        for idx, row in display_df.iterrows():
            cells = []

            for c in display_columns:
                v = row.get(c, "")

                if is_output_ineligible(row) and c in ineligible_blank_columns:
                    cells.append(
                        ft.DataCell(
                            ft.Container(
                                content=ft.Text(""),
                                alignment=(
                                    ft.alignment.center
                                    if should_center_align_column(c)
                                    else ft.alignment.center_left
                                ),
                            )
                        )
                    )
                    continue

                if c == "品番PG名":
                    def format_part_text(value) -> str:
                        if pd.isna(value):
                            return ""

                        s = str(value).strip()

                        if s == "" or s.lower() in ("nan", "none"):
                            return ""

                        return s

                    part_number = format_part_text(row.get("品番", ""))
                    pg_name = format_part_text(row.get("PG名", ""))

                    text_controls = []

                    if part_number:
                        text_controls.append(
                            ft.Text(
                                part_number,
                                no_wrap=True,
                            )
                        )

                    if pg_name:
                        text_controls.append(
                            ft.Text(
                                pg_name,
                                no_wrap=True,
                            )
                        )

                    if text_controls:
                        content = ft.Column(
                            controls=text_controls,
                            spacing=0,
                            tight=True,
                        )
                    else:
                        content = ft.Text("")

                    cells.append(
                        ft.DataCell(
                            ft.Container(
                                content=content,
                                alignment=ft.alignment.center_left,
                            )
                        )
                    )

                    continue

                if c == "プレビュー":
                    adopted_pdf_path = str(row.get("採用PDFパス", "") or "").strip()

                    if use_hover_preview:
                        content = ft.Text("")
                    elif adopted_pdf_path and not is_output_ineligible(row):
                        content = ft.IconButton(
                            icon=ft.Icons.VISIBILITY_OUTLINED,
                            tooltip="出力PDFをプレビュー",
                            on_click=lambda e, pdf_path=adopted_pdf_path: on_preview_adopted_pdf(pdf_path),
                        )
                    else:
                        content = ft.Text("")

                    cells.append(
                        ft.DataCell(
                            ft.Container(
                                content=content,
                                alignment=ft.alignment.center,
                            )
                        )
                    )

                    continue

                if c == "品番PG名":
                    def format_part_text(value) -> str:
                        if pd.isna(value):
                            return ""

                        s = str(value).strip()

                        if s == "" or s.lower() in ("nan", "none"):
                            return ""

                        return s

                    part_number = format_part_text(row.get("品番", ""))
                    pg_name = format_part_text(row.get("PG名", ""))

                    text_controls = []

                    if part_number:
                        text_controls.append(
                            ft.Text(
                                part_number,
                                no_wrap=True,
                            )
                        )

                    if pg_name:
                        text_controls.append(
                            ft.Text(
                                pg_name,
                                no_wrap=True,
                            )
                        )

                    if text_controls:
                        content = ft.Column(
                            controls=text_controls,
                            spacing=0,
                            tight=True,
                        )
                    else:
                        content = ft.Text("")

                    cells.append(
                        ft.DataCell(
                            ft.Container(
                                content=content,
                                alignment=ft.alignment.center_left,
                            )
                        )
                    )

                    continue

                if c == "候補確認":
                    if (not is_output_ineligible(row)) and row.get("候補状態") in ("複数候補", "手動採用"):
                        content = ft.ElevatedButton(
                            "選択",
                            tooltip="候補PDFを選択",
                            on_click=lambda e, row_index=idx: on_select_pdf_candidate(row_index),
                            style=ft.ButtonStyle(
                                bgcolor=ft.Colors.AMBER_500,
                                color=ft.Colors.BLACK,
                                padding=ft.padding.symmetric(horizontal=8, vertical=4),
                            ),
                        )
                    else:
                        content = ft.Text("")

                    cells.append(
                        ft.DataCell(
                            ft.Container(
                                content=content,
                                alignment=ft.alignment.center,
                            )
                        )
                    )

                    continue

                if c == "数量":
                    display_value = "該当" if is_zero_quantity(v) else ""

                elif c == "数量セル色":
                    display_value = "あり" if str(v).strip() else ""

                elif c == "出力可否":
                    raw_value = str(v or "").strip()

                    if raw_value == "可":
                        display_value = "対象"
                    elif raw_value == "不可":
                        display_value = "不可"
                    else:
                        display_value = ""

                elif c == "採用PDFパス":
                    adopted_pdf_path = str(row.get("採用PDFパス", "") or "").strip()

                    if adopted_pdf_path:
                        display_value = format_pdf_path_for_display(adopted_pdf_path)
                    else:
                        display_value = str(row.get("候補状態", "") or "").strip()

                elif c == "候補PDF数":
                    if pd.isna(v) or str(v).strip().lower() in ("", "nan", "none"):
                        display_value = ""
                    else:
                        display_value = str(v)

                else:
                    if pd.isna(v) or str(v).strip().lower() in ("nan", "none"):
                        display_value = ""
                    else:
                        display_value = str(v)

                if c == "出力対象":
                    adopted_pdf_path = str(row.get("採用PDFパス", "") or "").strip()

                    if is_output_ineligible(row) or not adopted_pdf_path:
                        cells.append(
                            ft.DataCell(
                                ft.Container(
                                    content=ft.Text(""),
                                    alignment=ft.alignment.center,
                                )
                            )
                        )
                    else:
                        checkbox = ft.Checkbox(
                            value=bool(v),
                            on_change=lambda e, row_index=idx: on_output_checkbox_change(e, row_index),
                        )
                        output_checkboxes[idx] = checkbox

                        cells.append(
                            ft.DataCell(
                                ft.Container(
                                    content=checkbox,
                                    alignment=ft.alignment.center,
                                )
                            )
                        )

                elif c == "検索用文字列":
                    if is_output_ineligible(row):
                        cells.append(
                            ft.DataCell(
                                ft.Container(
                                    width=110,
                                    padding=2,
                                    content=ft.Text(""),
                                )
                            )
                        )
                    else:
                        text_field = ft.TextField(
                            value=display_value,
                            dense=True,
                            text_size=14,
                            color=ft.Colors.BLACK,
                            content_padding=ft.padding.symmetric(horizontal=7, vertical=5),
                            border=ft.InputBorder.NONE,
                            bgcolor=ft.Colors.TRANSPARENT,
                        )
                        search_text_fields[idx] = text_field

                        cells.append(
                            ft.DataCell(
                                ft.Container(
                                    width=105,
                                    padding=2,
                                    bgcolor=ft.Colors.AMBER_50,
                                    border=ft.border.all(1, ft.Colors.AMBER_100),
                                    border_radius=6,
                                    content=text_field,
                                )
                            )
                        )

                else:
                    text_color = None

                    if c == "採用PDFパス":
                        adopted_pdf_path = str(row.get("採用PDFパス", "") or "").strip()
                        candidate_status = str(row.get("候補状態", "") or "").strip()

                        if adopted_pdf_path:
                            text_color = None
                        elif candidate_status == "未検出":
                            text_color = ft.Colors.RED_700
                        elif candidate_status == "複数候補":
                            text_color = ft.Colors.ORANGE_800
                        else:
                            text_color = ft.Colors.GREY_600

                    elif c == "出力可否":
                        if str(display_value).strip() == "不可":
                            text_color = ft.Colors.RED_700
                        else:
                            text_color = ft.Colors.GREEN_700

                    is_hover_pdf_column = c == "採用PDFパス"
                    adopted_pdf_path_for_hover = str(row.get("採用PDFパス", "") or "").strip()
                    enable_hover_preview_for_cell = (
                        use_hover_preview
                        and is_hover_pdf_column
                        and bool(adopted_pdf_path_for_hover)
                    )

                    normal_text_color = (
                        ft.Colors.BLUE_700
                        if enable_hover_preview_for_cell
                        else text_color
                    )

                    output_pdf_file_name_text = None

                    if c == "採用PDFパス" and adopted_pdf_path_for_hover:
                        folder_text, file_name = split_pdf_path_for_two_line_display(
                            adopted_pdf_path_for_hover
                        )

                        output_pdf_file_name_text = ft.Text(
                            file_name,
                            color=PDF_FILE_NAME_COLOR,
                            weight=ft.FontWeight.NORMAL,
                            no_wrap=False,
                            max_lines=2,
                            overflow=ft.TextOverflow.ELLIPSIS,
                        )

                        text_control = ft.Container(
                            width=OUTPUT_PDF_COLUMN_WIDTH,
                            content=ft.Column(
                                controls=[
                                    ft.Text(
                                        folder_text,
                                        color=ft.Colors.WHITE,
                                        weight=ft.FontWeight.NORMAL,
                                        no_wrap=True,
                                        overflow=ft.TextOverflow.ELLIPSIS,
                                    ),
                                    output_pdf_file_name_text,
                                ],
                                spacing=0,
                                tight=True,
                            ),
                        )
                    else:
                        text_control = ft.Text(
                            display_value,
                            color=normal_text_color,
                            weight=ft.FontWeight.NORMAL,
                            text_align=(
                                ft.TextAlign.CENTER
                                if should_center_align_column(c)
                                else ft.TextAlign.LEFT
                            ),
                        )

                    if enable_hover_preview_for_cell:
                        def handle_output_pdf_hover(
                            e,
                            pdf_path=adopted_pdf_path_for_hover,
                            file_name_text=output_pdf_file_name_text,
                        ):
                            if e.data == "true":
                                if file_name_text is not None:
                                    file_name_text.color = PDF_FILE_NAME_HOVER_COLOR

                                update_hover_preview(pdf_path)

                            else:
                                if file_name_text is not None:
                                    file_name_text.color = PDF_FILE_NAME_COLOR

                                clear_hover_preview()

                            page.update()

                        cell_content = ft.Container(
                            content=text_control,
                            alignment=ft.alignment.center_left,
                            on_hover=handle_output_pdf_hover,
                        )
                    else:
                        cell_content = ft.Container(
                            content=ft.SelectionArea(
                                content=text_control
                            ),
                            alignment=(
                                ft.alignment.center
                                if should_center_align_column(c)
                                else ft.alignment.center_left
                            ),
                        )

                    cells.append(ft.DataCell(cell_content))

            table.rows.append(ft.DataRow(cells=cells))

        update_hover_preview_visibility()

    # 🔽 抽出結果エリアの初期化関数を追加 🔽
    def reset_extract_view():
        nonlocal current_df

        current_df = pd.DataFrame()
        search_text_fields.clear()
        output_checkboxes.clear()

        table.rows = []
        table.columns = [ft.DataColumn(ft.Text("抽出結果"))]
        table_header.controls = []
        message.value = ""

        clear_hover_preview()
        update_hover_preview_visibility()

        page.update()

    def reset_excel_selection():
        """
        選択中のExcelファイルとシート選択を初期化する。
        """
        nonlocal selected_excel_path

        selected_excel_path = ""
        excel_file_field.value = ""

        sheet_dropdown.options = [
            ft.dropdown.Option(SHEET_PLACEHOLDER)
        ]
        sheet_dropdown.value = SHEET_PLACEHOLDER

        reset_extract_view()

    def on_sheet_change(e):
        # シート変更時に抽出結果を初期化
        reset_extract_view()

    sheet_dropdown.on_change = on_sheet_change

    file_picker = ft.FilePicker(on_result=lambda e: pick_excel_result(e))
    page.overlay.append(file_picker)

    def pick_excel_click(e):
        initial_dir = config.get("excel_folder", "")

        if not initial_dir or not os.path.isdir(initial_dir):
            initial_dir = os.getcwd()

        file_picker.pick_files(
            allowed_extensions=["xlsx", "xls"],
            initial_directory=initial_dir,
        )

    def pick_excel_result(e: ft.FilePickerResultEvent):
        nonlocal selected_excel_path  # sheet_dropdown は作り直さないので nonlocal 不要でもOK

        if not e.files:
            return

        # ★ 抽出エリアを初期化
        reset_extract_view()

        selected_excel_path = e.files[0].path
        excel_file_field.value = os.path.abspath(selected_excel_path)

        config["excel_folder"] = str(Path(selected_excel_path).parent)
        save_config()

        page.update()

        try:
            xls = pd.ExcelFile(selected_excel_path)
            sheet_names = xls.sheet_names

            # ✅ Dropdownを作り直さず、options/valueだけ更新する
            sheet_dropdown.options = [
                ft.dropdown.Option(SHEET_PLACEHOLDER),
                *[ft.dropdown.Option(name) for name in sheet_names]
            ]
            sheet_dropdown.value = SHEET_PLACEHOLDER
            sheet_dropdown.on_change = on_sheet_change  # 念のため（維持）
            sheet_dropdown.update()

            page.open(ft.SnackBar(ft.Text("シートを選択してください。")))
            page.update()

        except PermissionError:
            dialog = ft.AlertDialog(
                title=ft.Text("ファイル使用中"),
                content=ft.Text(
                    "Excelファイルを開けませんでした。\n\n"
                    "他のユーザー、または別のPCでこのExcelが開かれている可能性があります。\n"
                    "Excelファイルを閉じてから、もう一度選択してください。"
                ),
                actions_alignment="end",
            )

            dialog.actions = [
                ft.TextButton(
                    "OK",
                    on_click=lambda e: (
                        setattr(dialog, "open", False),
                        page.update(),
                    ),
                )
            ]

            page.open(dialog)
            page.update()


    # ---- ヘッダ検出 ----
    def get_merged_cells_info(excel_path, sheet_name):
        wb = load_workbook(excel_path, data_only=True)

        try:
            ws = wb[sheet_name]
            return list(ws.merged_cells.ranges)

        finally:
            wb.close()

    def get_vertically_merged_data_row_indices(merged_cells_info, header_row_index: int) -> set[int]:
        """
        ヘッダ行より下で、縦方向にまたがる結合セルを含むDataFrame上の行indexを返す。

        戻り値のindexは、pd.read_excel(..., header=header_row_index) で作った
        DataFrameのindexに対応する。

        横方向だけの結合セルは対象外。
        """
        merged_row_indices = set()

        # Excel上のデータ開始行。
        # header_row_index は 0始まり、openpyxlの行番号は1始まり。
        # pandasで header=header_row_index とした場合、
        # DataFrame index 0 は Excel上の header_row_index + 2 行目に対応する。
        data_start_excel_row = header_row_index + 2

        for crange in merged_cells_info:
            # 横結合だけなら対象外
            if crange.min_row == crange.max_row:
                continue

            # ヘッダ行以前の結合は対象外
            if crange.max_row < data_start_excel_row:
                continue

            start_excel_row = max(crange.min_row, data_start_excel_row)
            end_excel_row = crange.max_row

            for excel_row in range(start_excel_row, end_excel_row + 1):
                df_index = excel_row - data_start_excel_row
                if df_index >= 0:
                    merged_row_indices.add(df_index)

        return merged_row_indices

    def detect_header_row(df, merged_cells_info):
        merged_a_rows = set()
        for crange in merged_cells_info:
            if crange.min_col == 1:
                for r in range(crange.min_row - 1, crange.max_row):
                    merged_a_rows.add(r)
        for i, row in df.iterrows():
            a_val = str(row.iloc[0]).strip() if not pd.isna(row.iloc[0]) else ""
            filled = sum(1 for v in row if not pd.isna(v) and str(v).strip() != "")
            if i in merged_a_rows or not a_val:
                continue
            if filled >= 1:
                return i
        return 0

    # ---- 抽出処理 ----
    def on_extract_click(e):

        clear_hover_preview()

        # 1) Excel 未選択
        if not selected_excel_path:
            message.value = "Excelファイルを選択してください。"
            page.update()
            return

        # 2) シート未選択（プレースホルダ含む）
        sheet_val = sheet_dropdown.value
        if (not sheet_val) or (sheet_val == SHEET_PLACEHOLDER):
            message.value = "シートを選択してください。"
            page.update()
            return

        # 3) モードチェック
        if mode_dropdown.value != "構成部品表":
            message.value = "通常モードは未実装です。"
            page.update()
            return

        try:
            df_raw = pd.read_excel(selected_excel_path, sheet_name=sheet_dropdown.value, header=None)
            merged_info = get_merged_cells_info(selected_excel_path, sheet_dropdown.value)
            header_row_index = detect_header_row(df_raw, merged_info)
            detected_columns = detect_columns(selected_excel_path, sheet_dropdown.value, header_row_index)
            df = extract_data_from_excel(selected_excel_path, sheet_dropdown.value, detected_columns, header_row_index)

            vertical_merge_indices = get_vertically_merged_data_row_indices(
                merged_info,
                header_row_index,
            )

            def create_search_keyword(text):
                if pd.isna(text):
                    return ""

                s = str(text).strip()

                # 先頭に補足説明や記号が付いている場合、
                # 最初の半角英数字が出てくる位置まで削除する。
                #
                # 例:
                #   "補足説明 ABC123" -> "ABC123"
                #   "※ABC123"       -> "ABC123"
                #   "（参考）U12345" -> "U12345"
                match = re.search(r"[A-Za-z0-9]", s)
                if match:
                    s = s[match.start():]
                else:
                    return ""

                # 先頭が I/U/H/F + 数字 の場合は、先頭1文字を除去する
                if len(s) >= 2 and s[0].upper() in ["I", "U", "H", "F"] and s[1].isdigit():
                    s = s[1:]

                for pat in ["R/L", "L/R", "A/B"]:
                    idx = s.find(pat)
                    if idx != -1:
                        s = s[:idx]

                idx_slash = s.find("/")
                if idx_slash != -1:
                    s = s[:idx_slash]

                return s.strip()

            def resolve_pg_names_for_search(pg_series: pd.Series) -> dict:
                """
                検索文字列生成用にPG名を補正する。

                表示上のPG名は変更せず、
                PG名が ↑ / ↓ の場合だけ、検索用には直前のPG名を使う。
                """
                resolved = {}
                last_pg_name = ""

                for idx, value in pg_series.items():
                    if pd.isna(value):
                        resolved[idx] = ""
                        continue

                    s = str(value).strip()

                    if s in ("↑", "↓"):
                        resolved[idx] = last_pg_name
                        continue

                    resolved[idx] = s

                    if s and s.lower() not in ("nan", "none"):
                        last_pg_name = s

                return resolved

            # --- 品番 or PG名 のどちらか優先で生成（元値を保持）
            resolved_pg_names = resolve_pg_names_for_search(df["PG名"]) if "PG名" in df.columns else {}

            df["元検索文字列"] = df.apply(
                lambda row: (
                    create_search_keyword(row.get("品番"))
                    or create_search_keyword(resolved_pg_names.get(row.name, ""))
                ),
                axis=1
            )

            # --- 実際に検索に使う文字列（初期値は元検索文字列と同じ）
            df["検索用文字列"] = df["元検索文字列"]

            qty_colors = []
            if "数量" in detected_columns:
                qty_colors = get_quantity_colors(
                    selected_excel_path,
                    sheet_dropdown.value,
                    header_row_index,
                    detected_columns["数量"]
                )

                # ✅ df.index（元の行位置）で色リストを引く → フィルタ後でもズレない
                def color_by_df_index(i: int) -> str:
                    return qty_colors[i] if 0 <= i < len(qty_colors) else ""

                df["数量セル色"] = [color_by_df_index(i) for i in df.index]
            else:
                df["数量セル色"] = ""

            df["縦結合"] = [
                "あり" if i in vertical_merge_indices else ""
                for i in df.index
            ]

            def decide_output_eligibility(row):
                """
                Excel上の条件から、自動出力処理の対象にしてよいかを判定する。

                不可条件:
                - 数量が0
                - 数量セル色あり
                - 品番・PG名がともに空欄
                """
                def is_blank_value(value) -> bool:
                    if pd.isna(value):
                        return True

                    s = str(value).strip()

                    return s == "" or s.lower() in ("nan", "none")

                zero_qty = False

                if "数量" in df.columns:
                    zero_qty = is_zero_quantity(row.get("数量"))

                colored = bool(str(row.get("数量セル色", "") or "").strip())

                part_number_blank = is_blank_value(row.get("品番", ""))
                pg_name_blank = is_blank_value(row.get("PG名", ""))

                missing_identifier = part_number_blank and pg_name_blank

                if zero_qty or colored or missing_identifier:
                    return "不可"

                return "可"

            df["出力可否"] = df.apply(decide_output_eligibility, axis=1)

            # この時点では、まだ出力PDFが確定していないため、出力対象はすべてOFF。
            # PDF候補抽出後、出力可否が「可」かつ出力PDFが確定した行だけONにする。
            df["出力対象"] = False

            # 出力不可行は検索対象にしないため、検索文字列を空欄にする
            df.loc[df["出力可否"].eq("不可"), "検索用文字列"] = ""

            # PDF候補情報の初期列
            df["候補PDF数"] = ""
            df["先頭候補PDF"] = ""
            df["候補PDFパス一覧"] = [[] for _ in range(len(df))]
            df["採用PDFパス"] = ""
            df["候補状態"] = "未検索"

            df.loc[df["出力可否"].eq("不可"), "候補状態"] = "出力不可"

            # 採用PDF重複は、PDF候補抽出後に更新する
            df["採用PDF重複"] = ""

            if df.empty:
                message.value = "抽出結果がありません。"
                table.rows = []
                table_header.controls = []
            else:
                message.value = f"{len(df)}件のデータを抽出しました。"

                select_all_btn = ft.ElevatedButton(
                    "全選択",
                    on_click=lambda e: toggle_all(True),
                    style=BUTTON_STYLE_SECONDARY,
                )

                deselect_all_btn = ft.ElevatedButton(
                    "全解除",
                    on_click=lambda e: toggle_all(False),
                    style=BUTTON_STYLE_SECONDARY,
                )

                search_pdf_btn = ft.ElevatedButton(
                    "PDF候補抽出",
                    on_click=on_pdf_candidate_search,
                    style=BUTTON_STYLE_PRIMARY,
                )

                export_pdf_btn = ft.ElevatedButton(
                    "PDF出力実行",
                    on_click=on_pdf_export,
                    style=BUTTON_STYLE_SUCCESS,
                )

                table_header.controls = [
                    search_pdf_btn,
                    ft.Container(width=40),
                    select_all_btn,
                    deselect_all_btn,
                    export_pdf_btn,
                    ft.Container(width=240),
                ]

                render_table_from_df(df)


            page.update()
        except PermissionError:
            message.value = (
                "Excelファイルを開けませんでした。"
                "他のユーザー、または別のPCで開かれている可能性があります。"
            )
            page.open(
                ft.SnackBar(
                    ft.Text(
                        "Excelファイルを開けませんでした。"
                        "他のユーザー、または別のPCで開かれている可能性があります。"
                    )
                )
            )
            page.update()

        except Exception as ex:
            message.value = f"エラー: {ex}"
            page.update()

    def toggle_all(value: bool):
        nonlocal current_df

        for idx, checkbox in output_checkboxes.items():
            checkbox.value = value

            if current_df is not None and not current_df.empty and idx in current_df.index:
                current_df.at[idx, "出力対象"] = value

        render_table_from_df(current_df)
        page.update()

    def apply_output_target_defaults(df: pd.DataFrame) -> pd.DataFrame:
        """
        出力対象チェックの初期状態を更新する。

        条件:
        - 出力可否 が「可」
        - 採用PDFパス が空でない

        上記を満たす行だけ True にする。
        それ以外は False にする。
        """
        if df is None or df.empty:
            return df

        if "出力対象" not in df.columns:
            return df

        df = df.copy()

        if "出力可否" not in df.columns or "採用PDFパス" not in df.columns:
            df["出力対象"] = False
            return df

        eligible = df["出力可否"].fillna("").astype(str).str.strip().eq("可")
        has_output_pdf = df["採用PDFパス"].fillna("").astype(str).str.strip().ne("")

        df["出力対象"] = eligible & has_output_pdf

        return df

    def get_export_ready_df(df: pd.DataFrame) -> pd.DataFrame:
        """
        実際にPDF出力に回せる行だけを抽出する。

        条件:
        - 出力対象 が True
        - 採用PDFパス が空でない
        """
        if df is None or df.empty:
            return pd.DataFrame()

        if "出力対象" not in df.columns or "採用PDFパス" not in df.columns:
            return pd.DataFrame()

        output_target = df["出力対象"].fillna(False).astype(bool)
        adopted_pdf_exists = df["採用PDFパス"].fillna("").astype(str).str.strip().ne("")

        return df[output_target & adopted_pdf_exists].copy()

    def get_output_target_without_adopted_pdf_df(df: pd.DataFrame) -> pd.DataFrame:
        """
        出力対象にチェックが入っているが、採用PDFパスが空の行を抽出する。
        """
        if df is None or df.empty:
            return pd.DataFrame()

        if "出力対象" not in df.columns or "採用PDFパス" not in df.columns:
            return pd.DataFrame()

        output_target = df["出力対象"].fillna(False).astype(bool)
        adopted_pdf_missing = df["採用PDFパス"].fillna("").astype(str).str.strip().eq("")

        return df[output_target & adopted_pdf_missing].copy()

    def sanitize_filename_part(text: str) -> str:
        """
        Windowsのファイル名に使えない文字を置換する。
        """
        if text is None:
            return ""

        text = str(text).strip()

        invalid_chars = r'\/:*?"<>|'

        for ch in invalid_chars:
            text = text.replace(ch, "_")

        text = " ".join(text.split())

        return text

    def build_output_pdf_path(output_folder: str, excel_path: str, sheet_name: str) -> str:
        """
        出力先フォルダから、結合PDFの保存パスを作る。

        形式:
        Excelファイル名_シート名_YYYYMMDD_HHMM.pdf

        同名ファイルがある場合は _001, _002... を付ける。
        """
        if not output_folder or not output_folder.strip():
            raise ValueError("出力先フォルダが指定されていません。")

        output_dir = Path(output_folder)

        if not output_dir.exists():
            output_dir.mkdir(parents=True, exist_ok=True)

        if not output_dir.is_dir():
            raise ValueError("出力先フォルダが正しくありません。")

        now_text = datetime.now().strftime("%Y%m%d_%H%M%S")

        excel_stem = Path(excel_path).stem if excel_path else "Excel"
        sheet_text = sheet_name if sheet_name else "Sheet"

        excel_stem = sanitize_filename_part(excel_stem)
        sheet_text = sanitize_filename_part(sheet_text)

        base_name = f"{excel_stem}_{sheet_text}_{now_text}"

        candidate = output_dir / f"{base_name}.pdf"

        if not candidate.exists():
            return str(candidate)

        for i in range(1, 1000):
            candidate = output_dir / f"{base_name}_{i:03d}.pdf"
            if not candidate.exists():
                return str(candidate)

        raise ValueError("出力ファイル名を作成できませんでした。")

    def sync_table_state_to_current_df():
        """
        DataTable上の最新の検索用文字列・出力対象チェックを current_df に反映する。
        """
        nonlocal current_df

        if current_df is None or current_df.empty:
            return

        for idx, text_field in search_text_fields.items():
            if idx in current_df.index:
                current_df.at[idx, "検索用文字列"] = (text_field.value or "").strip()

        for idx, checkbox in output_checkboxes.items():
            if idx in current_df.index:
                current_df.at[idx, "出力対象"] = bool(checkbox.value)

    def on_pdf_candidate_search(e):
        nonlocal current_df

        if current_df is None or current_df.empty:
            message.value = "抽出結果がありません。"
            page.open(ft.SnackBar(ft.Text("抽出結果がありません。")))
            page.update()
            return

        pdf_root = search_folder_field.value.strip()

        if not pdf_root or not os.path.isdir(pdf_root):
            message.value = "検索先フォルダが正しくありません。"
            page.open(ft.SnackBar(ft.Text("検索先フォルダが正しくありません。")))
            page.update()
            return

        sync_table_state_to_current_df()

        message.value = "PDF一覧を取得中です..."
        page.update()

        def update_pdf_collect_progress(count: int):
            message.value = f"PDF一覧を取得中です... {count}件"
            page.update()

        pdf_files = collect_pdf_files(
            pdf_root,
            progress_callback=update_pdf_collect_progress
        )

        if not pdf_files:
            message.value = "検索先フォルダ配下にPDFが見つかりませんでした。"
            page.open(ft.SnackBar(ft.Text("検索先フォルダ配下にPDFが見つかりませんでした。")))
            page.update()
            return

        message.value = f"PDF検索用データを作成中です... {len(pdf_files)}件"
        page.update()

        pdf_index = build_pdf_search_index(pdf_files)

        search_target_count = len(current_df)

        message.value = f"PDF候補を抽出中です... 0 / {search_target_count}"
        page.update()

        # 各行ごとにPDF候補を抽出
        for pos, (idx, row) in enumerate(current_df.iterrows(), start=1):

            # 出力不可行はPDF検索対象外にする
            if is_output_ineligible(row):
                current_df.at[idx, "候補PDF数"] = ""
                current_df.at[idx, "先頭候補PDF"] = ""
                current_df.at[idx, "候補PDFパス一覧"] = []
                current_df.at[idx, "採用PDFパス"] = ""
                current_df.at[idx, "候補状態"] = "出力不可"
                current_df.at[idx, "出力対象"] = False

                if pos % 5 == 0 or pos == search_target_count:
                    message.value = f"PDF候補を抽出中です... {pos} / {search_target_count}"
                    page.update()

                continue

            search_text = row.get("検索用文字列", "")
            candidates = find_pdf_candidates_from_index(search_text, pdf_index)
            candidate_count = len(candidates)

            previous_status = str(row.get("候補状態", "") or "").strip()
            previous_adopted_pdf = str(row.get("採用PDFパス", "") or "").strip()

            current_df.at[idx, "候補PDF数"] = str(candidate_count) if candidate_count > 0 else ""
            current_df.at[idx, "先頭候補PDF"] = candidates[0] if candidates else ""
            current_df.at[idx, "候補PDFパス一覧"] = candidates

            if candidate_count == 0:
                current_df.at[idx, "採用PDFパス"] = ""
                current_df.at[idx, "候補状態"] = "未検出"

            elif candidate_count == 1:
                current_df.at[idx, "採用PDFパス"] = candidates[0]
                current_df.at[idx, "候補状態"] = "自動採用"

            else:
                # 複数候補の場合:
                # 以前に手動採用したPDFが今回の候補にも含まれていれば、その選択を保持する
                if previous_status == "手動採用" and previous_adopted_pdf in candidates:
                    current_df.at[idx, "採用PDFパス"] = previous_adopted_pdf
                    current_df.at[idx, "候補状態"] = "手動採用"
                else:
                    current_df.at[idx, "採用PDFパス"] = ""
                    current_df.at[idx, "候補状態"] = "複数候補"

            if pos % 5 == 0 or pos == search_target_count:
                message.value = f"PDF候補を抽出中です... {pos} / {search_target_count}"
                page.update()

        current_df = apply_output_target_defaults(current_df)
        current_df = update_adopted_pdf_duplicate_info(current_df)
        current_df = apply_duplicate_output_defaults(current_df)

        render_table_from_df(current_df)

        candidate_counts = pd.to_numeric(
            current_df["候補PDF数"],
            errors="coerce"
        ).fillna(0)

        matched_count = int((candidate_counts > 0).sum())
        selected_count = int(current_df["出力対象"].fillna(False).sum())
        adopted_count = int(
            current_df["採用PDFパス"]
            .fillna("")
            .astype(str)
            .str.strip()
            .ne("")
            .sum()
        )
        multiple_count = int((candidate_counts >= 2).sum())
        duplicate_adopted_pdf_count = int(
            current_df["採用PDF重複"]
            .fillna("")
            .astype(str)
            .str.strip()
            .ne("")
            .sum()
        )

        export_ready_df = get_export_ready_df(current_df)
        export_ready_count = len(export_ready_df)

        debug_print("===== PDF候補抽出結果 =====")
        for idx, row in current_df.iterrows():
            debug_print(
                f"[{idx}] "
                f"出力対象={row.get('出力対象', False)}, "
                f"検索用文字列={row.get('検索用文字列', '')!r}, "
                f"候補PDF数={row.get('候補PDF数', 0)}, "
                f"候補状態={row.get('候補状態', '')!r}, "
                f"先頭候補PDF={row.get('先頭候補PDF', '')!r}, "
                f"採用PDFパス={row.get('採用PDFパス', '')!r}"
            )

        debug_print("===== 複数候補行 =====")
        multiple_df = current_df[candidate_counts >= 2].copy()

        if multiple_df.empty:
            debug_print("複数候補の行はありません。")
        else:
            for idx, row in multiple_df.iterrows():
                debug_print(
                    f"[{idx}] "
                    f"検索用文字列={row.get('検索用文字列', '')!r}, "
                    f"候補PDF数={row.get('候補PDF数', 0)}"
                )

                candidates = row.get("候補PDFパス一覧", [])
                if not isinstance(candidates, list):
                    candidates = []

                for i, candidate_path in enumerate(candidates, start=1):
                    debug_print(f"  {i}. {candidate_path}")


        debug_print("===== 出力PDF 重複行 =====")
        duplicate_df = current_df[
            current_df["採用PDF重複"]
            .fillna("")
            .astype(str)
            .str.strip()
            .ne("")
        ].copy()

        if duplicate_df.empty:
            debug_print("検索用文字列が重複している行はありません。")
        else:
            for idx, row in duplicate_df.iterrows():
                debug_print(
                    f"[{idx}] "
                    f"検索用文字列={row.get('検索用文字列', '')!r}, "
                    f"重複={row.get('採用PDF重複', '')!r}, "
                    f"候補状態={row.get('候補状態', '')!r}, "
                    f"採用PDFパス={row.get('採用PDFパス', '')!r}"
                )

        debug_print("===== 出力可能行 =====")
        if export_ready_df.empty:
            debug_print("出力可能な行はありません。")
        else:
            for idx, row in export_ready_df.iterrows():
                debug_print(
                    f"[{idx}] "
                    f"検索用文字列={row.get('検索用文字列', '')!r}, "
                    f"採用PDFパス={row.get('採用PDFパス', '')!r}"
                )

        result_message = "PDF候補抽出が完了しました。"

        message.value = result_message

        page.open(
            ft.SnackBar(
                ft.Text(result_message)
            )
        )

        page.update()

    def on_pdf_export(e):
        nonlocal current_df

        if current_df is None or current_df.empty:
            message.value = "抽出結果がありません。"
            page.open(ft.SnackBar(ft.Text("抽出結果がありません。")))
            page.update()
            return

        sync_table_state_to_current_df()

        output_folder = output_folder_field.value.strip()

        if not output_folder:
            message.value = "出力先フォルダを指定してください。"
            page.open(ft.SnackBar(ft.Text("出力先フォルダを指定してください。")))
            page.update()
            return

        if os.path.exists(output_folder) and not os.path.isdir(output_folder):
            message.value = "出力先フォルダが正しくありません。"
            page.open(ft.SnackBar(ft.Text("出力先フォルダが正しくありません。")))
            page.update()
            return

        export_ready_df = get_export_ready_df(current_df)
        export_ready_count = len(export_ready_df)

        missing_pdf_df = get_output_target_without_adopted_pdf_df(current_df)
        missing_pdf_count = len(missing_pdf_df)

        if export_ready_df.empty:
            if missing_pdf_count > 0:
                result_message = (
                    "出力可能なPDFがありません。"
                    f"出力対象のうち {missing_pdf_count} 件は採用PDFが未確定です。"
                    "PDF候補抽出を実行するか、検索用文字列を見直してください。"
                )
            else:
                result_message = (
                    "出力可能なPDFがありません。"
                    "出力対象にチェックが入っている行、または採用PDFパスを確認してください。"
                )

            message.value = result_message
            page.open(ft.SnackBar(ft.Text(result_message)))
            page.update()
            return

        try:
            output_pdf_path = build_output_pdf_path(
                output_folder,
                selected_excel_path,
                sheet_dropdown.value,
            )

            pdf_paths = (
                export_ready_df["採用PDFパス"]
                .fillna("")
                .astype(str)
                .str.strip()
                .tolist()
            )

            pdf_paths = [p for p in pdf_paths if p]

            def update_export_progress(current: int, total: int, pdf_path: str):
                message.value = f"PDF出力中... {current} / {total}"
                page.update()

            message.value = f"PDF出力を開始します... 0 / {export_ready_count}"
            page.update()

            merge_pdfs(
                pdf_paths,
                output_pdf_path,
                progress_callback=update_export_progress,
            )

            result_message = f"PDF出力完了：{export_ready_count}件 / {os.path.basename(output_pdf_path)}"

            if missing_pdf_count > 0:
                result_message += f"（未確定除外：{missing_pdf_count}件）"

            debug_print("===== PDF出力 =====")
            debug_print(f"出力PDF: {output_pdf_path}")
            debug_print("===== 結合対象PDF =====")
            for p in pdf_paths:
                debug_print(p)

            if missing_pdf_count > 0:
                debug_print("===== 出力対象だが採用PDF未確定の行 =====")
                for idx, row in missing_pdf_df.iterrows():
                    debug_print(
                        f"[{idx}] "
                        f"検索用文字列={row.get('検索用文字列', '')!r}, "
                        f"候補PDF数={row.get('候補PDF数', 0)}, "
                        f"先頭候補PDF={row.get('先頭候補PDF', '')!r}"
                    )

        except Exception as ex:
            result_message = f"PDF出力エラー: {ex}"

        message.value = result_message

        page.open(
            ft.SnackBar(
                ft.Text(result_message)
            )
        )

        page.update()

    pick_excel_btn = ft.ElevatedButton(
        "Excelを選択",
        on_click=pick_excel_click,
        style=BUTTON_STYLE_PRIMARY,
    )

    extract_btn = ft.ElevatedButton(
        "抽出実行",
        on_click=on_extract_click,
        style=BUTTON_STYLE_PRIMARY,
    )

    def set_excel_ui_enabled(enabled: bool):
        excel_file_field.disabled = not enabled
        pick_excel_btn.disabled = not enabled
        sheet_dropdown.disabled = not enabled
        sheet_dropdown.opacity = 1.0 if enabled else 0.5
        extract_btn.disabled = not enabled

        if not enabled:
            reset_excel_selection()
            mode_notice.value = "通常モードでは構成部品表選択は無効です。"
        else:
            mode_notice.value = ""

        # 保険（効かない環境があるので）
        sheet_dropdown.update()
        page.update()

    # ---- レイアウト ----
    excel_file_field = ft.TextField(
        label="選択中のExcelファイル",
        expand=True,
        read_only=True,
    )
    folder_settings_row = ft.Row(
        [
            ft.Row(
                [
                    search_folder_field,
                    ft.ElevatedButton(
                        "フォルダを選択",
                        on_click=pick_search_folder,
                        style=BUTTON_STYLE_SECONDARY,
                    ),
                ],
                expand=True,
                spacing=6,
            ),
            ft.Row(
                [
                    output_folder_field,
                    ft.ElevatedButton(
                        "フォルダを選択",
                        on_click=pick_output_folder,
                        style=BUTTON_STYLE_SECONDARY,
                    ),
                ],
                expand=True,
                spacing=6,
            ),
        ],
        spacing=12,
    )

    excel_control_row = ft.Row(
        [
            ft.Container(
                content=excel_file_field,
                expand=2,
            ),
            pick_excel_btn,
            sheet_dropdown,
            extract_btn,
        ],
        alignment="center",
        vertical_alignment="center",
        spacing=10,
    )

    fixed_controls = ft.Container(
        content=ft.Column(
            [
                ft.Text("⚙️ 設定", size=20, weight="bold"),
                folder_settings_row,
                mode_row,

                ft.Container(height=4),
                ft.Divider(height=1),
                ft.Container(height=6),

                ft.Text("🔍 構成部品表選択", size=20, weight="bold"),
                mode_notice,
                excel_control_row,

                ft.Container(height=4),
                ft.Divider(height=1),

                table_header,
                message,
            ],
            spacing=8,
        ),
        padding=ft.padding.only(left=4, right=4, top=4, bottom=6),
    )

    table_area = ft.Container(
        content=ft.Column(
            [
                table,
            ],
            scroll="auto",
            expand=True,
        ),
    )

    result_area = ft.Row(
        controls=[
            table_area,            # 表は必要な幅だけ
            hover_preview_panel,   # 余った幅を全部使う
        ],
        expand=True,
        vertical_alignment=ft.CrossAxisAlignment.START,
    )

    layout = ft.Column(
        [
            fixed_controls,
            result_area,
        ],
        expand=True,
    )

    def update_hover_preview_visibility() -> bool:
        """
        画面幅と抽出結果の有無に応じて、ホバープレビュー欄の表示状態を切り替える。
        表示状態が変わった場合は True を返す。
        """
        nonlocal use_hover_preview

        page_width = page.width or 0
        has_results = current_df is not None and not current_df.empty

        new_use_hover_preview = (
            has_results
            and page_width >= HOVER_PREVIEW_MIN_WIDTH
        )

        debug_print(
            f"hover_preview: width={page_width}, "
            f"has_results={has_results}, "
            f"threshold={HOVER_PREVIEW_MIN_WIDTH}, "
            f"hover={new_use_hover_preview}"
        )

        changed = new_use_hover_preview != use_hover_preview

        use_hover_preview = new_use_hover_preview
        hover_preview_panel.visible = use_hover_preview

        if not use_hover_preview:
            clear_hover_preview()

        return changed

    def on_page_resize(e):

        changed = update_hover_preview_visibility()

        if changed and current_df is not None and not current_df.empty:
            render_table_from_df(current_df)

        page.update()

    page.on_resize = on_page_resize

    page.add(layout)

    update_hover_preview_visibility()
    update_mode_fields()


if __name__ == "__main__":
    ft.app(target=main)