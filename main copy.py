import flet as ft
import pandas as pd
import os
import json
import re
from openpyxl import load_workbook

CONFIG_FILE = "config.json"

def normalize(text: str):
    if not isinstance(text, str):
        return ""
    text = text.strip().replace("／", "/").replace("　", " ")
    return re.sub(r"\s+", "", text.lower())

def detect_columns(file_path: str, header_row_index: int, target_columns=None):
    if target_columns is None:
        target_columns = ["品番", "PG名", "品名", "数量"]

    df = pd.read_excel(file_path, header=header_row_index)
    detected = {}
    normalized_cols = [normalize(c) for c in df.columns]

    for target in target_columns:
        norm_target = normalize(target)
        for i, col in enumerate(normalized_cols):
            if norm_target in col:
                detected[target] = i
                break
    return detected

# ===== Excel抽出関数 =====
def extract_data_from_excel(file_path: str, detected_columns: dict, header_row_index: int):
    """
    Excelから指定列を抽出し、数量が0・空白のみを除外。
    品名は「品名」見出し列 + 直後の Unnamed 列群をまとめて連結（非空のみ結合）。
    """
    df = pd.read_excel(file_path, header=header_row_index)

    # --- まず検出済みキーの列だけ取り出し（無くても後で'品名'を上書きで作るのでOK） ---
    col_items = [col for col in detected_columns.keys() if col in df.columns]
    sub_df = df[col_items].copy()

    # ====== 品名列の堅牢な抽出・結合 ======
    cols = list(df.columns)

    def find_name_block_indices(cols):
        """'品名'を含む見出しの列を起点に、右側の Unnamed 列を連結対象として巻き取る"""
        for i, c in enumerate(cols):
            if "品名" in str(c):
                j = i + 1
                block = [i]
                # 右隣に 'Unnamed' や 空文字/None の見出しが続く限り巻き取る
                while j < len(cols):
                    cj = str(cols[j])
                    if cj.startswith("Unnamed") or cj.strip() == "" or cj.lower() == "none" or cj.lower() == "nan":
                        block.append(j)
                        j += 1
                    else:
                        break
                return block
        return []

    name_idx = find_name_block_indices(cols)

    if name_idx:
        name_df = df.iloc[:, name_idx].astype(str)
        # 行単位で非空のセルのみ連結（1列目空・2列目のみ値→2列目だけ残る）
        sub_df["品名"] = (
            name_df.apply(
                lambda row: " ".join(v.strip() for v in row if v.strip() and v.lower() != "nan"),
                axis=1
            ).str.strip()
        )
    else:
        # フォールバック：見出し名に「品名」を含む単独列がある場合
        name_like_cols = [c for c in df.columns if "品名" in str(c)]
        if len(name_like_cols) == 1:
            sub_df["品名"] = df[name_like_cols[0]].astype(str).fillna("").str.strip()
        # なければ品名は作らない（必要なら空列で用意してもOK）
        # else:
        #     sub_df["品名"] = ""

    # ====== 数量フィルタ（0と空白のみ除外、文字列は空でなければOK） ======
    quantity_col = next((c for c in df.columns if "数量" in str(c)), None)
    if quantity_col:
        def keep_row(x):
            if pd.isna(x):
                return False
            s = str(x).strip()
            if s == "":
                return False
            try:
                # 数値変換できた場合は 0 だけNG
                return float(s.replace(",", "")) != 0.0
            except ValueError:
                # 数値でなければ（例: '各1', '約3'）→ OK
                return True

        sub_df = sub_df[df[quantity_col].apply(keep_row)]

    # 品番は空でも除外しない（要件どおり）
    return sub_df

def main(page: ft.Page):
    page.title = "PDF検索ツール - Excel検索モード Step 2.2（項目名行検出対応）"
    page.scroll = "adaptive"

    # 仮のExcel情報（実際はフェーズ1で検出済み）
    # excel_path = "test_data.xlsx"
    # header_row_index = 2
    # detected_columns = {"品番": 2, "PG名": 4, "数量": 6}

    # ------------------------------
    # 設定ファイル
    # ------------------------------
    def load_config():
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        # return {"excel_folder": "", "pdf_folder": "", "output_folder": "", "search_mode": "Excel"}
        return {"excel_folder": "", "pdf_folder": "", "output_folder": "", "search_mode": "Excel", "target_columns": ["品番", "PG名"]}

    # def save_config():
    def save_config(e):
        config.update({
            "excel_folder": excel_folder_field.value,
            "pdf_folder": search_folder_field.value,
            "output_folder": output_folder_field.value,
            "search_mode": mode_dropdown.value,
            "target_columns": [x.strip() for x in target_col_field.value.split(",") if x.strip()]
        })
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(config, f, indent=2, ensure_ascii=False)

    config = load_config()

    # ------------------------------
    # UI 要素
    # ------------------------------
    excel_folder_field = ft.TextField(label="Excelフォルダ", value=config.get("excel_folder", ""), expand=True)
    search_folder_field = ft.TextField(label="検索先フォルダ（PDF）", value=config.get("pdf_folder", ""), expand=True)
    output_folder_field = ft.TextField(label="出力先フォルダ", value=config.get("output_folder", ""), expand=True)
    target_col_field = ft.TextField(
        label="抽出対象列（カンマ区切り）",
        value=", ".join(config.get("target_columns", [])),
        width=400
    )
    
    # def save_target_columns(e):
    #     config["target_columns"] = [x.strip() for x in target_col_field.value.split(",") if x.strip()]
    #     save_config(config)
    #     page.snack_bar = ft.SnackBar(ft.Text("抽出対象列の設定を保存しました。"))
    #     page.snack_bar.open = True
    #     page.update()

    mode_dropdown = ft.Dropdown(
        label="検索モード",
        options=[ft.dropdown.Option("Excel"), ft.dropdown.Option("通常")],
        value=config.get("search_mode", "Excel"),
    )

    selected_excel_path = ""
    sheet_dropdown = ft.Dropdown(label="シート選択", width=300)
    column_dropdown = ft.Dropdown(label="列選択（項目名行に基づく）", width=400)
    extracted_keywords = ft.ListView(expand=True, spacing=5)

    # ------------------------------
    # Excel ヘルパー関数
    # ------------------------------
    def get_merged_cells_info(excel_path, sheet_name):
        """openpyxl で結合セル情報を取得"""
        wb = load_workbook(excel_path, data_only=True)
        ws = wb[sheet_name]
        return ws.merged_cells.ranges

    def detect_header_row(df, merged_cells_info):
        """
        項目名行を自動検出：
        ・A列に値があり
        ・結合セルを含まず
        ・少なくとも 2 列以上に値がある行
        """
        merged_rows = set()
        for crange in merged_cells_info:
            for r in range(crange.min_row - 1, crange.max_row):
                merged_rows.add(r)

        for i, row in df.iterrows():
            if i in merged_rows:
                continue
            a_val = str(row.iloc[0]).strip() if not pd.isna(row.iloc[0]) else ""
            if not a_val:
                continue
            filled = sum(1 for v in row if not pd.isna(v) and str(v).strip() != "")
            if filled >= 2:
                return i
        return 0

    # ------------------------------
    # イベント処理
    # ------------------------------
    def pick_excel_result(e: ft.FilePickerResultEvent):
        nonlocal selected_excel_path, sheet_dropdown, column_dropdown

        if not e.files:
            return

        selected_excel_path = e.files[0].path
        excel_folder_field.value = os.path.dirname(selected_excel_path)

        # ✅ Excel選択時：シート・列選択のコントロールを完全再生成
        sheet_dropdown = ft.Dropdown(
            label="シート選択",
            width=300,
            on_change=on_sheet_selected,  # ✅ ユーザー選択時に自動読み込み
        )
        column_dropdown = ft.Dropdown(label="列選択（項目名行に基づく）", width=400)
        extracted_keywords.controls.clear()

        # 再生成してUIに再配置
        layout.controls[2] = ft.Row(
            [sheet_dropdown],  # ✅ シート読み込みボタン削除
            alignment="center",
        )
        layout.controls[3] = ft.Row(
            [column_dropdown, ft.ElevatedButton("列を読み込む", on_click=on_column_selected)],
            alignment="center",
        )

        page.update()

        # ✅ Excelのシート一覧を取得（初期値はNone、ユーザー選択を待つ）
        try:
            xls = pd.ExcelFile(selected_excel_path)
            sheet_dropdown.options = [ft.dropdown.Option(name) for name in xls.sheet_names]
            sheet_dropdown.value = None  # ✅ 自動選択しない
            page.snack_bar = ft.SnackBar(ft.Text(f"{len(xls.sheet_names)} シートを読み込みました（選択してください）"))
            page.snack_bar.open = True
            page.update()
        except Exception as ex:
            page.dialog = ft.AlertDialog(title=ft.Text("Excel読込エラー"), content=ft.Text(str(ex)))
            page.dialog.open = True
            page.update()


    def on_sheet_selected(e):
        """シート選択時：列Dropdownを再生成"""
        nonlocal column_dropdown

        # ✅ 列Dropdownを完全に再生成してUIに反映
        column_dropdown = ft.Dropdown(label="列選択（項目名行に基づく）", width=400)
        extracted_keywords.controls.clear()
        layout.controls[3] = ft.Row(
            [column_dropdown, ft.ElevatedButton("列を読み込む", on_click=on_column_selected)],
            alignment="center",
        )
        page.update()

        if not selected_excel_path or not sheet_dropdown.value:
            return

        try:
            df = pd.read_excel(selected_excel_path, sheet_name=sheet_dropdown.value, header=None)
            merged_info = get_merged_cells_info(selected_excel_path, sheet_dropdown.value)
            header_row = detect_header_row(df, merged_info)
            headers = df.iloc[header_row].tolist()
            df = df.iloc[header_row + 1:]
            column_dropdown.options = [ft.dropdown.Option(str(h)) for h in headers if str(h).strip() != ""]
            column_dropdown.value = headers[0] if headers else None

            page.snack_bar = ft.SnackBar(ft.Text(f"項目名行 {header_row + 1} 行目を検出しました"))
            page.snack_bar.open = True
            page.update()
        except Exception as ex:
            page.dialog = ft.AlertDialog(title=ft.Text("シート読込エラー"), content=ft.Text(str(ex)))
            page.dialog.open = True
            page.update()


    def on_column_selected(e):
        """列選択 → 文字列抽出"""
        if not selected_excel_path or not column_dropdown.value:
            return
        try:
            # Step 1: まず header_row を再検出（検出ロジックを再利用）
            df_raw = pd.read_excel(selected_excel_path, sheet_name=sheet_dropdown.value, header=None)
            merged_info = get_merged_cells_info(selected_excel_path, sheet_dropdown.value)
            header_row = detect_header_row(df_raw, merged_info)

            # Step 2: 正しい項目名行を header として再読み込み
            df = pd.read_excel(
                selected_excel_path,
                sheet_name=sheet_dropdown.value,
                header=header_row
            )

            # Step 3: 列名を正規化
            df.columns = [str(c).strip().replace("\n", "").replace("\r", "") for c in df.columns]
            selected_col = str(column_dropdown.value).strip().replace("\n", "").replace("\r", "")

            print("▼ 実際の列名一覧:", df.columns.tolist())
            print("▼ 選択中の列名:", selected_col)

            # Step 4: 列一致チェック
            matching_cols = [c for c in df.columns if c == selected_col]
            if not matching_cols:
                raise ValueError(f"列 '{selected_col}' が見つかりません。")

            col = matching_cols[0]
            print("▼ 一致した列:", col)

            # Step 5: 検索文字列抽出
            keywords = []
            for val in df[col].dropna():
                s = str(val).strip()
                if not s:
                    continue
                if re.match(r"^[IUHF]", s) and len(s) > 1 and s[1].isdigit():
                    s = s[1:]
                keywords.append(s)

            # Step 6: 表示更新
            extracted_keywords.controls.clear()
            for kw in keywords:
                extracted_keywords.controls.append(ft.Text(kw))
            page.snack_bar = ft.SnackBar(ft.Text(f"{len(keywords)} 件の検索文字列を抽出しました"))
            page.snack_bar.open = True
            page.update()

        except Exception as ex:
            page.dialog = ft.AlertDialog(title=ft.Text("列読込エラー"), content=ft.Text(str(ex)))
            page.dialog.open = True
            page.update()


    # ------------------------------
    # FilePicker
    # ------------------------------
    file_picker = ft.FilePicker(on_result=pick_excel_result)
    page.overlay.append(file_picker)

    def pick_excel_click(e):
        file_picker.pick_files(allowed_extensions=["xlsx", "xls"])

    # UI要素
    message = ft.Text("")
    table = ft.DataTable(
        columns=[ft.DataColumn(ft.Text("項目"))],
        rows=[]
    )

    def on_extract_click(e):
        if not selected_excel_path:
            message.value = "Excelファイルを先に選択してください。"
            page.update()
            return
        try:
            # 自動検出処理
            df_raw = pd.read_excel(selected_excel_path, sheet_name=sheet_dropdown.value, header=None)
            merged_info = get_merged_cells_info(selected_excel_path, sheet_dropdown.value)
            header_row_index = detect_header_row(df_raw, merged_info)
            detected_columns = detect_columns(selected_excel_path, header_row_index)
            
            df = extract_data_from_excel(selected_excel_path, detected_columns, header_row_index)
            if df.empty:
                message.value = "抽出結果がありません。"
                table.columns = [ft.DataColumn(ft.Text("項目"))]  # リセットしてもOK
                table.rows = []
            else:
                message.value = f"{len(df)}件のデータを抽出しました。"
                table.columns = [ft.DataColumn(ft.Text(c)) for c in df.columns]
                table.rows = [
                    ft.DataRow(cells=[ft.DataCell(ft.Text(str(v))) for v in row])
                    for row in df.values.tolist()
                ]
            page.update()
        except Exception as ex:
            message.value = f"エラー: {ex}"
            page.update()

    # ボタン追加
    extract_button = ft.ElevatedButton("Excelデータ抽出", on_click=on_extract_click)
    
    # ------------------------------
    # レイアウト
    # ------------------------------
    layout = ft.Column([
        ft.Text("🔍 Excel検索モード", size=20, weight=ft.FontWeight.BOLD),
        ft.Row([excel_folder_field, ft.ElevatedButton("Excelを選択", on_click=pick_excel_click)], alignment="center"),
        ft.Row([sheet_dropdown], alignment="center"),  # ✅ ボタン削除済
        ft.Row([column_dropdown, ft.ElevatedButton("列を読み込む", on_click=on_column_selected)], alignment="center"),
        ft.Divider(),
        ft.Text("📄 抽出された検索文字列:"),
        extracted_keywords,

        ft.Divider(),
        ft.Text("⚙️ 抽出対象列の設定:"),
        ft.Row(
            # [target_col_field, ft.ElevatedButton("保存", on_click=save_target_columns)],
            # [target_col_field, ft.ElevatedButton("設定を保存", on_click=lambda e: save_config())],
            [target_col_field, ft.ElevatedButton("設定を保存", on_click=save_config)],
            alignment="center"
        ),
        ft.Text("Excel–PDF結合アプリ（フェーズ2）", size=20, weight="bold"),
        extract_button,
        message,
        table

    ], expand=True, scroll="auto")

    page.add(layout)


ft.app(target=main)
